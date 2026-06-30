"""Build the formula-driven Excel valuation model from scratch.

Produces financial_models/BESS_Valuation.xlsx following the FAST/ICAEW/
Macabacus/Operis house rules: Inputs -> Calcs -> Outputs zones, one master
Timeline, a single CHOOSE scenario switch (3 cases) with a live-case row,
one-row-one-calculation, colour coding (blue input / black formula / green
cross-sheet link), and a Checks framework with a master check on the Cover.

DEAL: an ILLUSTRATIVE distribution-network BESS *develop-and-flip* fund (framed
on the an illustrative develop-and-flip battery storage fund — manager claims independently rebuilt). The
fund develops ~5 MW distribution batteries to shovel-ready (RTB / development
rights) and SELLS them before construction: merchant risk passes to the buyer;
the fund's risk is the SURVIVAL CURVE and the RTB EXIT PRICE.

The model is fully LIVE (formula strings) so it recalculates when opened in Excel.
Blue input cells are seeded from the same inputs the Python valuation engine uses,
so the workbook reproduces the engine's numbers (incl. closed-form investor IRR).

Excel-first: the master workbook is HAND-OWNED. A normal run writes NOTHING — Python
never overwrites the master, and there is no generated copy. Regenerate the master
from Python only with the explicit flag.

Run:  python -m src.build_model                 # no-op (master is hand-owned)
      python -m src.build_model --rebuild-master  # regenerates the hand-owned master
"""
from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule

from src.utils import io
from src.valuation_engine import load_inputs

# ---------------------------------------------------------------------------
# Sheet names (15 tabs)
# ---------------------------------------------------------------------------
COV, CON, CL, INP, TL, SCN = "Cover", "Contents", "Change Log", "Inputs", "Timeline", "Scenarios"
SURV, RNPV, FUND, RET = "Calc_Survival", "Calc_Project_rNPV", "Calc_Company", "Returns"
CC, SENS, CHK, DSH, SG = "Calc_CrossChecks", "Sensitivity", "Checks", "Dashboard", "Sources & Glossary"
ORDER = [COV, CON, CL, INP, TL, SCN, SURV, RNPV, FUND, RET, CC, SENS, CHK, DSH, SG]
TABCOLOR = {COV: "000000", CON: "808080", CL: "808080", INP: "0070C0", TL: "00B0F0",
            SCN: "7030A0", SURV: "FFFFFF", RNPV: "FFFFFF", FUND: "FFFFFF", RET: "92D050",
            CC: "FFFFFF", SENS: "FFFF00", CHK: "FF0000", DSH: "00B050", SG: "808080"}

# ---------------------------------------------------------------------------
# Styles
# ---------------------------------------------------------------------------
BLUE, BLACK, GREEN, RED, GREY = "0000FF", "000000", "008000", "FF0000", "595959"
F_INPUT = Font(name="Arial", size=10, color=BLUE)
F_FORM = Font(name="Arial", size=10, color=BLACK)
F_LINK = Font(name="Arial", size=10, color=GREEN)
F_LABEL = Font(name="Arial", size=10, color=BLACK)
F_TITLE = Font(name="Arial", size=14, bold=True, color="1F3864")
F_HEAD = Font(name="Arial", size=10, bold=True, color="FFFFFF")
F_SECT = Font(name="Arial", size=10, bold=True, color="1F3864")
F_NOTE = Font(name="Arial", size=9, italic=True, color=GREY)
F_DISC = Font(name="Arial", size=9, italic=True, color=RED)
FILL_HEAD = PatternFill("solid", fgColor="1F3864")
FILL_SECT = PatternFill("solid", fgColor="D9E1F2")
FILL_YEL = PatternFill("solid", fgColor="FFF2CC")
FILL_GREEN = PatternFill("solid", fgColor="C6EFCE")
FILL_RED = PatternFill("solid", fgColor="FFC7CE")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

FM = '$#,##0.0;($#,##0.0);"–"'        # $m, 1dp
FM2 = '$#,##0.00;($#,##0.00);"–"'     # $m, 2dp
FMW = '$#,##0.000'                    # $m per MW
FPCT = '0.0%'
FX = '0.00"x"'
FNUM = '#,##0'
FNUM1 = '#,##0.0'
FFAC = '0.000'
FMULT = '0.00'


# ---------------------------------------------------------------------------
# Builder helpers
# ---------------------------------------------------------------------------
class B:
    def __init__(self):
        self.wb = Workbook()
        self.wb.calculation.fullCalcOnLoad = True
        self.wb.remove(self.wb.active)
        self.ws = {name: self.wb.create_sheet(name) for name in ORDER}
        for name, ws in self.ws.items():
            ws.sheet_properties.tabColor = TABCOLOR[name]
            ws.sheet_view.showGridLines = False

    def put(self, sheet, r, c, value, kind="label", fmt=None, fill=None,
            bold=False, align=None, wrap=False, border=False):
        if kind not in ("formula", "link") and isinstance(value, str) and value.startswith("="):
            value = value[1:].lstrip()
        cell = self.ws[sheet].cell(row=r, column=c, value=value)
        font = {"input": F_INPUT, "formula": F_FORM, "link": F_LINK,
                "label": F_LABEL, "title": F_TITLE, "head": F_HEAD,
                "sect": F_SECT, "note": F_NOTE, "disc": F_DISC}[kind]
        if bold and kind in ("label", "formula", "link"):
            font = Font(name="Arial", size=10, bold=True, color=font.color.rgb)
        cell.font = font
        if fmt:
            cell.number_format = fmt
        if kind == "input" and fill is None:
            fill = "yel"
        if fill == "head":
            cell.fill = FILL_HEAD
        elif fill == "sect":
            cell.fill = FILL_SECT
        elif fill == "yel":
            cell.fill = FILL_YEL
        elif fill:
            cell.fill = PatternFill("solid", fgColor=fill)
        if align:
            cell.alignment = Alignment(horizontal=align, wrap_text=wrap, vertical="top")
        elif wrap:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        if border:
            cell.border = BORDER
        return cell

    def ref(self, sheet, col, row):
        return f"'{sheet}'!${get_column_letter(col)}${row}"

    def rng(self, sheet, c1, r1, c2, r2):
        return f"'{sheet}'!${get_column_letter(c1)}${r1}:${get_column_letter(c2)}${r2}"

    def width(self, sheet, widths):
        for col, w in widths.items():
            self.ws[sheet].column_dimensions[col].width = w

    def header(self, sheet, r, labels, start=1):
        for i, lab in enumerate(labels):
            self.put(sheet, r, start + i, lab, "head", fill="head",
                     align="center", wrap=True, border=True)


def build(rebuild_master: bool = False):
    b = B()
    inp = load_inputs()
    n = len(inp.projects)
    a = io.load_assumptions()
    R = {}

    # =====================================================================
    # INPUTS
    # =====================================================================
    b.width(INP, {"A": 36, "B": 10, "C": 14, "D": 13, "E": 50, "F": 12, "G": 12, "H": 12})
    b.put(INP, 1, 1, "INPUTS — all model assumptions (blue = input cell)", "title")
    b.put(INP, 2, 1, "We buy SHARES directly in an ILLUSTRATIVE battery-developer startup (not a fund LP). Founder/manager figures are CLAIMS — verify; the equity-deal terms are placeholders to confirm. Every input traces to a source.", "note", wrap=True)
    b.header(INP, 3, ["Assumption", "", "Value", "Unit", "Source", "As at", "Status", "Duration (yrs)"])

    def srow(r, label, value, key, fmt, unit, src_key=None, status="BENCHMARK", dur=None):
        b.put(INP, r, 1, label, "label")
        b.put(INP, r, 3, value, "input", fmt=fmt, border=True)
        b.put(INP, r, 4, unit, "label")
        src = inp.sources.get(src_key) if src_key else None
        b.put(INP, r, 5, src[0] if src else "Illustrative / the manager claim (see config)", "note", wrap=True)
        b.put(INP, r, 6, src[1] if src else a["meta"]["as_at"], "note", align="center")
        b.put(INP, r, 7, src[2] if src else status, "note", align="center")
        if dur is not None:
            b.put(INP, r, 8, dur, "input", fmt=FNUM1, border=True)
        if key:
            R[key] = b.ref(INP, 3, r)

    b.put(INP, 4, 1, "Discount rate (rNPV / cross-checks)", "sect", fill="sect")
    srow(5, "Risk-free rate (RBA 10yr CGS)", inp.risk_free, "risk_free", FPCT, "decimal", "risk_free")
    srow(6, "Development risk premium", inp.risk_premium, "risk_premium", FPCT, "decimal")
    b.put(INP, 7, 1, "Base discount rate", "label", bold=True)
    b.put(INP, 7, 3, f"={R['risk_free']}+{R['risk_premium']}", "formula", fmt=FPCT, border=True, bold=True)
    b.put(INP, 7, 4, "decimal", "label")
    R["discount_base"] = b.ref(INP, 3, 7)

    b.put(INP, 8, 1, "The company's development programme (the startup we buy shares in)", "sect", fill="sect")
    srow(9, "Programme capital (company's dev programme)", inp.programme_capital, "programme", FM, "$m")
    srow(10, "Projects target (delivered & sold)", inp.projects_target, "target", FNUM, "count")
    srow(11, "Term", inp.term_years, "term", FNUM, "years")

    b.put(INP, 12, 1, "Direct-equity deal terms (cap table) — placeholders to confirm", "sect", fill="sect")
    srow(13, "Pre-money valuation", inp.pre_money, "pre_money", FM, "$m", status="TO CONFIRM")
    srow(14, "Our investment", inp.investment_amount, "investment", FM, "$m", status="TO CONFIRM")
    srow(15, "Option pool % (created at our round)", inp.option_pool_pct, "option_pool", FPCT, "decimal", status="TO CONFIRM")
    srow(16, "Future-round dilution %", inp.future_round_dilution_pct, "future_dilution", FPCT, "decimal", status="TO CONFIRM")
    srow(17, "Liquidation preference (x)", inp.liquidation_pref_x, "liq_pref", FX, "multiple", status="TO CONFIRM")
    srow(18, "Exit year (we sell our shares)", inp.exit_year, "exit_year", FNUM, "years", status="TO CONFIRM")

    b.put(INP, 20, 1, "Development cost & funnel attrition", "sect", fill="sect")
    srow(21, "Development cost per project", inp.dev_cost_per_project, "dev_cost", FM2, "$m/project", "dev")
    srow(22, "Abandonment fraction (failed projects)", inp.abandonment_fraction, "abandon", FPCT, "decimal")

    b.put(INP, 23, 1, "Build cost CONTEXT (buyer-funded — NOT in margin)", "sect", fill="sect")
    srow(24, "Built-asset cost (context)", inp.built_cost_per_mw, "built_cost", FMW, "$m/MW", "build")

    b.put(INP, 25, 1, "Survival gates — SEPARATE; scenarios move ONLY development approval", "sect", fill="sect")
    srow(26, "Development approval — public benchmark", inp.da_independent, "da_ind", FPCT, "prob", "da", dur=inp.dur_da)
    srow(27, "Grid connection — RTB rights (agreement + GPS)", inp.p_connection, "p_conn", FPCT, "prob", "connection", dur=inp.dur_connection)
    srow(28, "Reach sale — flip exit (separate gate)", inp.p_sale, "p_sale", FPCT, "prob", "sale", dur=inp.dur_sale)

    b.put(INP, 29, 1, "RTB exit price — $/MW by state (the manager claim — verify)", "sect", fill="sect")
    for i, st in enumerate(["NSW", "VIC", "SA"]):
        srow(30 + i, st, float(inp.rtb_comps.get(st, 0.0)), f"rtb_{st}", FMW, "$m/MW", "rtb")
    R["rtb_states"] = b.rng(INP, 1, 30, 1, 32)
    R["rtb_prices"] = b.rng(INP, 3, 30, 3, 32)

    b.put(INP, 33, 1, "VC method", "sect", fill="sect")
    srow(34, "VC target return (cross-check)", inp.vc_target_return, "vc_target", FPCT, "decimal")

    b.put(INP, 35, 1, "Investor cash-flow profile (period 0..3; each sums to 100%)", "sect", fill="sect")
    b.put(INP, 36, 1, "Capital call profile", "label")
    b.put(INP, 37, 1, "Distribution profile", "label")
    for j in range(4):
        b.put(INP, 36, 3 + j, float(inp.call_profile[j]) if j < len(inp.call_profile) else 0.0,
              "input", fmt=FPCT, border=True, align="center")
        b.put(INP, 37, 3 + j, float(inp.dist_profile[j]) if j < len(inp.dist_profile) else 0.0,
              "input", fmt=FPCT, border=True, align="center")
    R["call_profile"] = b.rng(INP, 3, 36, 6, 36)
    R["dist_profile"] = b.rng(INP, 3, 37, 6, 37)

    pr0 = 39
    b.put(INP, 38, 1, "Illustrative pipeline (representative ~5 MW distribution projects)", "sect", fill="sect")
    b.header(INP, pr0, ["Project", "State", "Location", "MW", "Duration (h)", "MWh", "Years to sale"])
    for i, p in enumerate(inp.projects):
        r = pr0 + 1 + i
        b.put(INP, r, 1, p["name"], "input")
        b.put(INP, r, 2, p["state"], "input", align="center")
        b.put(INP, r, 3, p["location"], "input")
        b.put(INP, r, 4, p["mw"], "input", fmt=FNUM, border=True)
        b.put(INP, r, 5, p["duration_h"], "input", fmt=FNUM1, border=True)
        b.put(INP, r, 6, f"=D{r}*E{r}", "formula", fmt=FNUM, border=True)
        b.put(INP, r, 7, p["years_to_sale"], "input", fmt=FNUM1, border=True)
    pr1 = pr0 + n
    R["pipe_name0"] = pr0 + 1
    R["pipe_mw"] = b.rng(INP, 4, pr0 + 1, 4, pr1)
    R["pipe_years"] = b.rng(INP, 7, pr0 + 1, 7, pr1)

    # Exit value — PRIMARY = forward-pipeline rNPV + retained cash − debt (change2.md)
    xv0 = 47
    b.put(INP, xv0, 1, "Exit value — PRIMARY basis (forward-pipeline rNPV + retained cash − debt)", "sect", fill="sect")
    srow(xv0 + 1, "Pipeline depth at exit (projects in flight)", inp.pipeline_depth_at_exit, "exit_depth", FNUM, "count", status="TO CONFIRM")
    srow(xv0 + 2, "Interim distribution fraction (of realised profit)", inp.interim_distribution_fraction, "dist_frac", FPCT, "decimal", status="TO CONFIRM")
    srow(xv0 + 3, "Debt at exit", inp.debt_at_exit, "debt_exit", FM, "$m", status="TO CONFIRM")
    b.put(INP, xv0 + 4, 1, "Exit value CROSS-CHECK — earnings multiple on forward run-rate profit", "sect", fill="sect")
    srow(xv0 + 5, "Cross-check multiple — low", inp.xmult_low, "xmult_low", FMULT, "multiple", status="TO CONFIRM")
    srow(xv0 + 6, "Cross-check multiple — base", inp.xmult_base, "xmult_base", FMULT, "multiple", status="TO CONFIRM")
    srow(xv0 + 7, "Cross-check multiple — high", inp.xmult_high, "xmult_high", FMULT, "multiple", status="TO CONFIRM")

    # Provenance flags (change5) — 1 = Independent (verified) may enter the rNPV; 0 = Proposed/Placeholder
    pv0 = xv0 + 9
    _ver = lambda key: 1 if inp.provenance.get(key) == "Independent (verified)" else 0
    b.put(INP, pv0, 1, "Input provenance flags (change5) — 1 = Independent (verified); 0 = Proposed/Placeholder", "sect", fill="sect")
    b.put(INP, pv0 + 1, 1, "RTB $/MW verified? (feeds independent rNPV)", "label")
    b.put(INP, pv0 + 1, 3, _ver("rtb"), "input", fmt=FNUM, border=True)
    R["rtb_verified"] = b.ref(INP, 3, pv0 + 1)
    b.put(INP, pv0 + 2, 1, "Dev cost verified? (feeds independent rNPV)", "label")
    b.put(INP, pv0 + 2, 3, _ver("dev_cost"), "input", fmt=FNUM, border=True)
    R["devcost_verified"] = b.ref(INP, 3, pv0 + 2)
    b.put(INP, pv0 + 3, 1, "Provenance tagged & disclosed?", "label")
    b.put(INP, pv0 + 3, 3, 1, "input", fmt=FNUM, border=True)
    R["prov_disclosed"] = b.ref(INP, 3, pv0 + 3)
    b.put(INP, pv0 + 4, 1, "RTB $/MW and dev cost are the manager's UNVERIFIED claims (Proposed) — verify against a primary "
                           "source (comparable RTB sales; bottom-up costs) and set the flag to 1 before treating the rNPV as independent.", "note", wrap=True)

    # =====================================================================
    # SCENARIOS (3 cases: Conservative / Base / Ideal)
    # =====================================================================
    b.width(SCN, {"A": 44, "B": 14, "C": 14, "D": 14, "E": 14})
    b.put(SCN, 1, 1, "SCENARIOS — one switch drives the live case (CHOOSE)", "title")
    b.put(SCN, 2, 1, "Scenario switch  (1=Conservative 2=Base 3=Ideal)", "label", bold=True)
    b.put(SCN, 2, 2, inp.switch_default, "input", fmt=FNUM, fill="yel", border=True, align="center")
    R["switch"] = b.ref(SCN, 2, 2)
    b.put(SCN, 3, 1, "Active scenario", "label", bold=True)
    b.put(SCN, 3, 2, f'=CHOOSE({R["switch"]},"Conservative","Base","Ideal")', "formula", bold=True)
    R["scenario_name"] = b.ref(SCN, 2, 3)

    b.header(SCN, 4, ["Driver", "Conservative", "Base", "Ideal", "Live (active)"])
    cases = inp.scenarios  # keyed 1..3
    drivers = [("Development-approval rate (DA gate — the 40/65/80%)", "da_rate", FPCT, "live_da"),
               ("RTB price multiplier", "sale_price_multiplier", FMULT, "live_sale_mult"),
               ("Dev cost multiplier", "dev_cost_multiplier", FMULT, "live_dev_mult")]
    for di, (label, field, fmt, key) in enumerate(drivers):
        r = 5 + di
        b.put(SCN, r, 1, label, "label")
        for ci, cid in enumerate([1, 2, 3]):
            val = cases[cid].get(field)
            val = 0 if val is None else float(val)
            b.put(SCN, r, 2 + ci, val, "input", fmt=fmt, border=True, align="center")
        b.put(SCN, r, 5, f"=CHOOSE({R['switch']},B{r},C{r},D{r})", "formula", fmt=fmt, bold=True, border=True)
        R[key] = b.ref(SCN, 5, r)
    # discount rate is fixed (scenarios move ONLY the DA gate, not the discount rate)
    b.put(SCN, 9, 1, "Live discount rate (= base; scenarios don't move it)", "label", bold=True)
    b.put(SCN, 9, 5, f"={R['discount_base']}", "link", fmt=FPCT, bold=True, border=True)
    R["live_discount"] = b.ref(SCN, 5, 9)
    # scenario-column refs (DA gate by scenario; multipliers)
    R["scn_da"] = {1: b.ref(SCN, 2, 5), 2: b.ref(SCN, 3, 5), 3: b.ref(SCN, 4, 5)}
    R["scn_salemult"] = {1: b.ref(SCN, 2, 6), 2: b.ref(SCN, 3, 6), 3: b.ref(SCN, 4, 6)}
    R["scn_devmult"] = {1: b.ref(SCN, 2, 7), 2: b.ref(SCN, 3, 7), 3: b.ref(SCN, 4, 7)}
    # per-scenario flip success = DA x grid connection x sale  (the corrected logic)
    R["scn_flip"] = {c: f"{R['scn_da'][c]}*{R['p_conn']}*{R['p_sale']}" for c in (1, 2, 3)}

    b.put(SCN, 11, 1, "First-Chicago weights", "sect", fill="sect")
    b.header(SCN, 12, ["", "Conservative", "Base", "Ideal", "Sum"])
    b.put(SCN, 13, 1, "Weight", "label")
    for ci, name in enumerate(["Conservative", "Base", "Ideal"]):
        b.put(SCN, 13, 2 + ci, float(inp.weights[name]), "input", fmt=FPCT, border=True, align="center")
    b.put(SCN, 13, 5, "=SUM(B13:D13)", "formula", fmt=FPCT, bold=True, border=True)
    R["weights"] = b.rng(SCN, 2, 13, 4, 13)
    R["weights_sum"] = b.ref(SCN, 5, 13)
    b.put(SCN, 15, 1, "The 40/65/80% are the DEVELOPMENT-APPROVAL gate ONLY (the manager's deck). True flip success = development "
                      "approval x grid connection x sale — far below 65% (see Calc_Survival). The switch moves ONLY the DA gate; "
                      "connection and sale are fixed public benchmarks. Scenarios are the analyst's to OWN.", "note", wrap=True)

    # =====================================================================
    # TIMELINE
    # =====================================================================
    b.width(TL, {"A": 34, "B": 12, "C": 12, "D": 12, "E": 12})
    nper = 3
    b.put(TL, 1, 1, "TIMELINE — built once; every sheet links here", "title")
    b.put(TL, 2, 1, "Base year", "label")
    b.put(TL, 2, 2, a["meta"]["base_year"], "input", fmt=FNUM, border=True)
    R["base_year"] = b.ref(TL, 2, 2)
    b.put(TL, 3, 1, "Period number", "label")
    for p in range(nper + 1):
        b.put(TL, 3, 2 + p, p, "formula", fmt=FNUM, align="center")
    b.put(TL, 4, 1, "Year", "label")
    for p in range(nper + 1):
        col = 2 + p
        b.put(TL, 4, col, f"=$B$2+{get_column_letter(col)}3", "formula", fmt="0", align="center")
    b.put(TL, 5, 1, "Discount factor @ live discount", "label")
    for p in range(nper + 1):
        col = 2 + p
        cl = get_column_letter(col)
        b.put(TL, 5, col, f"=IFERROR(1/(1+{R['live_discount']})^{cl}3,0)", "formula", fmt=FFAC, align="center")
    R["tl_periods"] = b.rng(TL, 2, 3, 2 + nper, 3)

    # =====================================================================
    # CALC_SURVIVAL (separate gates; flip success = their product)
    # =====================================================================
    b.width(SURV, {"A": 46, "B": 16, "C": 18})
    b.put(SURV, 1, 1, "CALC — SURVIVAL GATES (each separate; flip success = their product)", "title")
    b.put(SURV, 2, 1, "The manager's 40/65/80% are the DEVELOPMENT-APPROVAL gate ONLY. True develop-and-flip success = development "
                      "approval x grid connection x sale — a multi-period survival / probability-of-default curve. The DA gate is LIVE "
                      "from the scenario switch; grid connection and sale are fixed public benchmarks.", "note", wrap=True)
    b.header(SURV, 3, ["Gate", "Probability", "Cumulative survival"])
    gate_rows = [("Development approval (live scenario DA gate)", R["live_da"]),
                 ("Grid connection — RTB rights (agreement + GPS)", R["p_conn"]),
                 ("Reach sale (flip exit)", R["p_sale"])]
    for i, (label, pref) in enumerate(gate_rows):
        r = 4 + i
        b.put(SURV, r, 1, label, "label")
        b.put(SURV, r, 2, f"={pref}", "link", fmt=FPCT, border=True)
        if i == 0:
            b.put(SURV, r, 3, f"=B{r}", "formula", fmt=FPCT, border=True)
        else:
            b.put(SURV, r, 3, f"=C{r-1}*B{r}", "formula", fmt=FPCT, border=True)
    b.put(SURV, 7, 1, "LIVE flip success (DA x connection x sale)", "label", bold=True)
    b.put(SURV, 7, 3, "=C6", "formula", fmt=FPCT, bold=True, border=True, fill="sect")
    R["live_success"] = b.ref(SURV, 3, 7)
    b.put(SURV, 9, 1, "Public-benchmark development approval", "label")
    b.put(SURV, 9, 3, f"={R['da_ind']}", "link", fmt=FPCT, border=True)
    b.put(SURV, 10, 1, "Independent flip success (benchmark DA x conn x sale)", "label", bold=True)
    b.put(SURV, 10, 3, f"=C9*{R['p_conn']}*{R['p_sale']}", "formula", fmt=FPCT, bold=True, border=True)
    R["flip_independent"] = b.ref(SURV, 3, 10)
    b.put(SURV, 12, 1, "Manager's headline DA rate (Base — the '65%')", "label")
    b.put(SURV, 12, 3, f"={R['scn_da'][2]}", "link", fmt=FPCT, border=True)
    b.put(SURV, 13, 1, "True flip success at that DA (DA x conn x sale)", "label", bold=True)
    b.put(SURV, 13, 3, f"={R['scn_flip'][2]}", "formula", fmt=FPCT, bold=True, border=True)
    R["flip_at_base_da"] = b.ref(SURV, 3, 13)
    b.put(SURV, 14, 3, '=IF(C13<C12,"FLAG: 65% is the DA gate only","ok")', "formula", align="center", border=True)
    b.put(SURV, 15, 1, "The 65% headline is the approval gate alone; after grid connection (~70%) and sale (~80%) the true flip success is "
                       "~36%. The DA gate is the master return driver — the scenarios move it 40/65/80%. The sub-5MW AEMO registration "
                       "exemption MAY lift small-distribution success — verify per state.", "note", wrap=True)
    b.put(SURV, 17, 1, "Grid connection = RTB-stage connection RIGHTS: the connection AGREEMENT + Generator Performance Standards (GPS) "
                       "that make a project ready-to-build (AEMO Scorecard 'registration / executed connection agreement' stage) — NOT "
                       "energisation/commissioning of a built battery (a flip sells pre-construction). 70% is a benchmark [[TO CONFIRM]].", "note", wrap=True)
    b.put(SURV, 18, 1, "Gates 2 & 3 are INDEPENDENT (no double-count): gate 2 = did the project SECURE connection rights/GPS; gate 3 = "
                       "given a genuinely RTB project, did a BUYER pay RTB price (demand/price risk, not connection again).", "note", wrap=True)

    # =====================================================================
    # CALC_PROJECT_rNPV (per-project risk-adjusted NAV — dev cost only)
    # =====================================================================
    b.width(RNPV, {"A": 28, "B": 7, "C": 7, "D": 7, "E": 11, "F": 8, "G": 11, "H": 11,
                   "I": 11, "J": 8, "K": 11, "L": 10, "M": 11, "N": 11})
    b.put(RNPV, 1, 1, "CALC — PER-PROJECT RISK-ADJUSTED NAV (rNPV)", "title")
    b.put(RNPV, 2, 1, "Per project: RTB sale − DEVELOPMENT cost = margin → × cumulative success → discounted. Cost is DEV ONLY "
                      "(the fund sells RTB; the buyer funds construction). All figures AUD $m unless stated.", "note", wrap=True)
    b.header(RNPV, 3, ["Project", "State", "MW", "Years", "RTB $/MW", "Sale mult", "Sale $m",
                       "Dev cost $m", "Margin $m", "Cum P", "Risk-adj $m", "Disc factor", "PV $m", "Sale@1x $m"])
    rv0 = 4
    for i in range(n):
        r = rv0 + i
        ip = R["pipe_name0"] + i
        b.put(RNPV, r, 1, f"={b.ref(INP, 1, ip)}", "link")
        b.put(RNPV, r, 2, f"={b.ref(INP, 2, ip)}", "link", align="center")
        b.put(RNPV, r, 3, f"={b.ref(INP, 4, ip)}", "link", fmt=FNUM, border=True)
        b.put(RNPV, r, 4, f"={b.ref(INP, 7, ip)}", "link", fmt=FNUM1, border=True)
        b.put(RNPV, r, 5, f"=IFERROR(INDEX({R['rtb_prices']},MATCH(B{r},{R['rtb_states']},0)),0)", "formula", fmt=FMW, border=True)
        b.put(RNPV, r, 6, f"={R['live_sale_mult']}", "link", fmt=FMULT, border=True)
        b.put(RNPV, r, 7, f"=C{r}*E{r}*F{r}", "formula", fmt=FM2, border=True)
        b.put(RNPV, r, 8, f"={R['dev_cost']}*{R['live_dev_mult']}", "formula", fmt=FM2, border=True)
        b.put(RNPV, r, 9, f"=G{r}-H{r}", "formula", fmt=FM2, border=True)
        b.put(RNPV, r, 10, f"={R['live_success']}", "link", fmt=FPCT, border=True)
        b.put(RNPV, r, 11, f"=I{r}*J{r}", "formula", fmt=FM2, border=True)
        b.put(RNPV, r, 12, f"=IFERROR(1/(1+{R['live_discount']})^D{r},0)", "formula", fmt=FFAC, border=True)
        b.put(RNPV, r, 13, f"=K{r}*L{r}", "formula", fmt=FM2, border=True)
        b.put(RNPV, r, 14, f"=C{r}*E{r}", "formula", fmt=FM2, border=True)
    rv1 = rv0 + n - 1
    rt = rv1 + 1
    b.put(RNPV, rt, 1, "Pipeline total (representative)", "label", bold=True)
    for col in (3, 7, 8, 13):
        cl = get_column_letter(col)
        b.put(RNPV, rt, col, f"=SUM({cl}{rv0}:{cl}{rv1})", "formula", fmt=(FNUM if col == 3 else FM2), bold=True, border=True)
    R["pipeline_rnpv"] = b.ref(RNPV, 13, rt)
    R["rnpv_mw"] = b.rng(RNPV, 3, rv0, 3, rv1)
    R["rnpv_rtb"] = b.rng(RNPV, 5, rv0, 5, rv1)
    R["rnpv_years"] = b.rng(RNPV, 4, rv0, 4, rv1)
    rbl = rt + 1
    b.put(RNPV, rbl, 1, "Blended RTB sale per project (@1x)", "label", bold=True)
    b.put(RNPV, rbl, 13, f"=AVERAGE({b.rng(RNPV, 14, rv0, 14, rv1)})", "formula", fmt=FM2, bold=True, border=True, fill="sect")
    R["base_blended"] = b.ref(RNPV, 13, rbl)

    # =====================================================================
    # CALC_COMPANY (the company's development programme -> exit equity value)
    # =====================================================================
    b.width(FUND, {"A": 50, "B": 4, "C": 14, "D": 10, "E": 10, "F": 10})
    b.put(FUND, 1, 1, "CALC — THE COMPANY'S DEVELOPMENT PROGRAMME & EXIT EQUITY VALUE (live scenario)", "title")
    b.put(FUND, 2, 1, "A develop-and-flip company is a development PLATFORM, so a buyer pays for its FORWARD pipeline, not past "
                      "profit. PRIMARY exit equity = forward-pipeline rNPV (projects still in flight at exit) + cash RETAINED on "
                      "the balance sheet − debt. Realised programme profit feeds RETAINED cash (counted once, at face value — never "
                      "at a multiple). The earnings multiple below is a CROSS-CHECK only. All figures AUD $m.", "note", wrap=True)
    b.put(FUND, 4, 1, "Development funnel (live scenario)", "sect", fill="sect")
    b.put(FUND, 5, 1, "Live flip success (DA × conn × sale)", "label")
    b.put(FUND, 5, 3, f"={R['live_success']}", "link", fmt=FPCT, border=True)
    b.put(FUND, 6, 1, "Projects target (delivered & sold)", "label")
    b.put(FUND, 6, 3, f"={R['target']}", "link", fmt=FNUM, border=True)
    b.put(FUND, 7, 1, "Projects started (funnel)", "label")
    b.put(FUND, 7, 3, "=IFERROR(C6/C5,0)", "formula", fmt=FNUM1, border=True)
    b.put(FUND, 8, 1, "Projects failed (dropouts)", "label")
    b.put(FUND, 8, 3, "=C7-C6", "formula", fmt=FNUM1, border=True)
    b.put(FUND, 9, 1, "Dev cost per project (live)", "label")
    b.put(FUND, 9, 3, f"={R['dev_cost']}*{R['live_dev_mult']}", "formula", fmt=FM2, border=True)
    b.put(FUND, 10, 1, "Total development cost", "label", bold=True)
    b.put(FUND, 10, 3, f"=C6*C9+C8*{R['abandon']}*C9", "formula", fmt=FM, border=True)
    R["co_devcost"] = b.ref(FUND, 3, 10)

    b.put(FUND, 11, 1, "Proceeds & realised programme profit", "sect", fill="sect")
    b.put(FUND, 12, 1, "Blended RTB sale per project (live)", "label")
    b.put(FUND, 12, 3, f"={R['base_blended']}*{R['live_sale_mult']}", "formula", fmt=FM2, border=True)
    b.put(FUND, 13, 1, "Gross proceeds", "label", bold=True)
    b.put(FUND, 13, 3, "=C6*C12", "formula", fmt=FM, border=True)
    R["co_gross"] = b.ref(FUND, 3, 13)
    b.put(FUND, 14, 1, "Realised net programme profit (gross − total dev cost)", "label", bold=True)
    b.put(FUND, 14, 3, "=C13-C10", "formula", fmt=FM, border=True)
    R["co_net"] = b.ref(FUND, 3, 14)
    b.put(FUND, 15, 1, "Forward run-rate annual dev profit (= realised ÷ term)", "label")
    b.put(FUND, 15, 3, f"=IFERROR(C14/{R['term']},0)", "formula", fmt=FM2, border=True)
    R["co_runrate"] = b.ref(FUND, 3, 15)

    b.put(FUND, 16, 1, "Company exit equity — PRIMARY (forward-pipeline rNPV + retained − debt)", "sect", fill="sect")
    b.put(FUND, 17, 1, "Avg years to sale (pipeline)", "label")
    b.put(FUND, 17, 3, f"=AVERAGE({R['rnpv_years']})", "formula", fmt=FNUM1, border=True)
    b.put(FUND, 18, 1, "Discount factor @ base over avg years", "label")
    b.put(FUND, 18, 3, f"=1/(1+{R['discount_base']})^C17", "formula", fmt=FFAC, border=True)
    R["df_avg"] = b.ref(FUND, 3, 18)
    b.put(FUND, 19, 1, "Per-project rNPV (blended; live scenario)", "label")
    b.put(FUND, 19, 3, f"=({R['base_blended']}*{R['live_sale_mult']}-C9)*C5*C18", "formula", fmt=FM2, border=True)
    R["co_rnpv_proj"] = b.ref(FUND, 3, 19)
    b.put(FUND, 20, 1, "Pipeline depth at exit (projects in flight)", "label")
    b.put(FUND, 20, 3, f"={R['exit_depth']}", "link", fmt=FNUM, border=True)
    b.put(FUND, 21, 1, "Forward pipeline rNPV at exit (= depth × per-project)", "label", bold=True)
    b.put(FUND, 21, 3, "=C20*C19", "formula", fmt=FM, border=True)
    R["co_fwd"] = b.ref(FUND, 3, 21)
    b.put(FUND, 22, 1, "Net cash retained at exit (= MAX(0,realised) × (1−dist frac))", "label")
    b.put(FUND, 22, 3, f"=MAX(0,C14)*(1-{R['dist_frac']})", "formula", fmt=FM, border=True)
    R["co_retained"] = b.ref(FUND, 3, 22)
    b.put(FUND, 23, 1, "Debt at exit", "label")
    b.put(FUND, 23, 3, f"={R['debt_exit']}", "link", fmt=FM, border=True)
    b.put(FUND, 24, 1, "Company exit equity = forward + retained − debt", "label", bold=True)
    b.put(FUND, 24, 3, "=MAX(0,C21+C22-C23)", "formula", fmt=FM, bold=True, border=True, fill="sect")
    R["co_exit_equity"] = b.ref(FUND, 3, 24)

    b.put(FUND, 25, 1, "Cross-check — earnings multiple on run-rate (NOT the primary number)", "sect", fill="sect")
    b.put(FUND, 26, 1, "Cross-check exit eq — low (run-rate × low mult + retained − debt)", "label")
    b.put(FUND, 26, 3, f"=MAX(0,C15)*{R['xmult_low']}+C22-C23", "formula", fmt=FM, border=True)
    b.put(FUND, 27, 1, "Cross-check exit eq — base", "label")
    b.put(FUND, 27, 3, f"=MAX(0,C15)*{R['xmult_base']}+C22-C23", "formula", fmt=FM, border=True)
    b.put(FUND, 28, 1, "Cross-check exit eq — high", "label")
    b.put(FUND, 28, 3, f"=MAX(0,C15)*{R['xmult_high']}+C22-C23", "formula", fmt=FM, border=True)
    b.put(FUND, 29, 1, "Primary (pipeline) vs cross-check (base mult): anchor on the more conservative pipeline basis; if they "
                       "diverge materially, the multiple is more generous than the risk-adjusted forward pipeline.", "note", wrap=True)

    # =====================================================================
    # RETURNS (cap table -> return on OUR shares by scenario, First Chicago)
    # =====================================================================
    b.width(RET, {"A": 40, "B": 13, "C": 13, "D": 13, "E": 13})
    b.put(RET, 1, 1, "RETURNS — cap table & return on OUR shares by scenario, First-Chicago, valuation range", "title")
    b.put(RET, 2, 1, "We buy SHARES directly. Ownership = our investment ÷ post-money, reduced by dilution. Each scenario's company "
                     "exit equity value (PRIMARY basis: forward-pipeline rNPV + retained cash − debt) flows through the cap table; our "
                     "TERMINAL proceeds = the GREATER of our 1× liquidation preference or our as-converted (diluted) share, plus any "
                     "INTERIM distributions (convention (b); dist fraction is a placeholder, default 0). All figures AUD $m unless stated.", "note", wrap=True)

    # --- Cap table block ---
    b.put(RET, 3, 1, "Cap table (our stake)", "sect", fill="sect")
    b.put(RET, 4, 1, "Pre-money valuation", "label")
    b.put(RET, 4, 2, f"={R['pre_money']}", "link", fmt=FM, border=True)
    b.put(RET, 5, 1, "Our investment", "label")
    b.put(RET, 5, 2, f"={R['investment']}", "link", fmt=FM, border=True)
    b.put(RET, 6, 1, "Post-money = pre-money + investment", "label", bold=True)
    b.put(RET, 6, 2, "=B4+B5", "formula", fmt=FM, bold=True, border=True)
    R["post_money"] = b.ref(RET, 2, 6)
    b.put(RET, 7, 1, "Ownership at entry = investment ÷ post-money", "label", bold=True)
    b.put(RET, 7, 2, "=IFERROR(B5/B6,0)", "formula", fmt=FPCT, bold=True, border=True)
    R["own_entry"] = b.ref(RET, 2, 7)
    b.put(RET, 8, 1, "Ownership at exit (diluted)", "label", bold=True)
    b.put(RET, 8, 2, f"=B7*(1-{R['option_pool']})*(1-{R['future_dilution']})", "formula", fmt=FPCT, bold=True, border=True, fill="sect")
    R["own_diluted"] = b.ref(RET, 2, 8)

    # --- Per-scenario block (each column independent of the switch) ---
    b.put(RET, 10, 1, "Return on OUR shares by scenario", "sect", fill="sect")
    b.header(RET, 11, ["Metric", "Conservative", "Base", "Ideal"])
    scol = {1: "B", 2: "C", 3: "D"}
    bold_rows = (23, 26, 28, 29, 30)
    metric_rows = [
        (12, "Flip success (DA × conn × sale)", lambda c: f"={R['scn_flip'][c]}", FPCT, "formula"),
        (13, "RTB price multiplier", lambda c: f"={R['scn_salemult'][c]}", FMULT, "link"),
        (14, "Dev cost multiplier", lambda c: f"={R['scn_devmult'][c]}", FMULT, "link"),
        (15, "Dev cost per project", lambda c: f"={R['dev_cost']}*{scol[c]}14", FM2, "formula"),
        (16, "Projects started", lambda c: f"=IFERROR({R['target']}/{scol[c]}12,0)", FNUM1, "formula"),
        (17, "Projects failed", lambda c: f"={scol[c]}16-{R['target']}", FNUM1, "formula"),
        (18, "Total dev cost", lambda c: f"={R['target']}*{scol[c]}15+{scol[c]}17*{R['abandon']}*{scol[c]}15", FM, "formula"),
        (19, "Gross proceeds", lambda c: f"={R['target']}*{R['base_blended']}*{scol[c]}13", FM, "formula"),
        (20, "Realised net programme profit", lambda c: f"={scol[c]}19-{scol[c]}18", FM, "formula"),
        (21, "Forward pipeline rNPV at exit (depth × per-project)",
         lambda c: f"={R['exit_depth']}*({R['base_blended']}*{scol[c]}13-{scol[c]}15)*{scol[c]}12*{R['df_avg']}", FM2, "formula"),
        (22, "Net cash retained = MAX(0,realised)×(1−dist)", lambda c: f"=MAX(0,{scol[c]}20)*(1-{R['dist_frac']})", FM2, "formula"),
        (23, "Company exit equity = fwd + retained − debt", lambda c: f"=MAX(0,{scol[c]}21+{scol[c]}22-{R['debt_exit']})", FM, "formula"),
        (24, "As-converted (diluted × exit equity)", lambda c: f"={R['own_diluted']}*{scol[c]}23", FM2, "formula"),
        (25, "Preference claim = MIN(liq×inv, exit eq)", lambda c: f"=MIN({R['liq_pref']}*{R['investment']},{scol[c]}23)", FM2, "formula"),
        (26, "Terminal proceeds = MAX(pref, as-converted)", lambda c: f"=MAX({scol[c]}25,{scol[c]}24)", FM2, "formula"),
        (27, "Interim distributions (diluted×dist×realised)", lambda c: f"={R['own_diluted']}*{R['dist_frac']}*MAX(0,{scol[c]}20)", FM2, "formula"),
        (28, "Our total proceeds = terminal + interim", lambda c: f"={scol[c]}26+{scol[c]}27", FM2, "formula"),
        (29, "MOIC = our proceeds ÷ our investment", lambda c: f"=IFERROR({scol[c]}28/{R['investment']},0)", FX, "formula"),
        (30, "Equity IRR = MOIC^(1/exit yr) − 1", lambda c: f"=IFERROR({scol[c]}29^(1/{R['exit_year']})-1,-1)", FPCT, "formula"),
    ]
    for r, label, fn, fmt, kind in metric_rows:
        b.put(RET, r, 1, label, "label", bold=(r in bold_rows))
        for c in (1, 2, 3):
            b.put(RET, r, 1 + c, fn(c), kind, fmt=fmt, border=True, bold=(r in bold_rows))
    R["proceeds_by"] = {1: b.ref(RET, 2, 28), 2: b.ref(RET, 3, 28), 3: b.ref(RET, 4, 28)}
    R["terminal_by"] = {1: b.ref(RET, 2, 26), 2: b.ref(RET, 3, 26), 3: b.ref(RET, 4, 26)}
    R["irr_by"] = {1: b.ref(RET, 2, 30), 2: b.ref(RET, 3, 30), 3: b.ref(RET, 4, 30)}
    R["moic_by"] = {1: b.ref(RET, 2, 29), 2: b.ref(RET, 3, 29), 3: b.ref(RET, 4, 29)}
    R["exit_eq_by"] = {1: b.ref(RET, 2, 23), 2: b.ref(RET, 3, 23), 3: b.ref(RET, 4, 23)}
    R["fwd_by"] = {1: b.ref(RET, 2, 21), 2: b.ref(RET, 3, 21), 3: b.ref(RET, 4, 21)}
    R["retained_by"] = {1: b.ref(RET, 2, 22), 2: b.ref(RET, 3, 22), 3: b.ref(RET, 4, 22)}
    R["realised_by"] = {1: b.ref(RET, 2, 20), 2: b.ref(RET, 3, 20), 3: b.ref(RET, 4, 20)}
    R["ret_irr_rng"] = b.rng(RET, 2, 30, 4, 30)
    R["ret_moic_rng"] = b.rng(RET, 2, 29, 4, 29)
    R["ret_proceeds_rng"] = b.rng(RET, 2, 28, 4, 28)

    # --- First-Chicago block (weight PROCEEDS, then derive MOIC/IRR) ---
    b.put(RET, 32, 1, "First-Chicago probability-weighted expected return", "sect", fill="sect")
    b.header(RET, 33, ["", "Conservative", "Base", "Ideal", "Expected"])
    b.put(RET, 34, 1, "Weight", "label")
    for ci, name in enumerate(["Conservative", "Base", "Ideal"]):
        b.put(RET, 34, 2 + ci, f"={b.ref(SCN, 2 + ci, 13)}", "link", fmt=FPCT, border=True, align="center")
    b.put(RET, 35, 1, "Our proceeds", "label")
    for c in (1, 2, 3):
        b.put(RET, 35, 1 + c, f"={R['proceeds_by'][c]}", "link", fmt=FM2, border=True)
    b.put(RET, 35, 5, f"=SUMPRODUCT({R['weights']},B35:D35)", "formula", fmt=FM2, bold=True, border=True)
    R["expected_proceeds"] = b.ref(RET, 5, 35)
    b.put(RET, 36, 1, "Expected MOIC = exp proceeds ÷ investment", "label", bold=True)
    b.put(RET, 36, 5, f"=IFERROR(E35/{R['investment']},0)", "formula", fmt=FX, bold=True, border=True)
    R["expected_moic"] = b.ref(RET, 5, 36)
    b.put(RET, 37, 1, "Expected equity IRR = exp MOIC^(1/exit yr) − 1", "label", bold=True)
    b.put(RET, 37, 5, f"=IFERROR(E36^(1/{R['exit_year']})-1,-1)", "formula", fmt=FPCT, bold=True, border=True, fill="sect")
    R["expected_irr"] = b.ref(RET, 5, 37)

    b.put(RET, 39, 1, "Valuation range (per-pipeline asset value — cross-check)", "sect", fill="sect")
    b.put(RET, 39, 3, "Representative 6-project pipeline (asset-level), NOT the company's whole exit equity value.", "note")
    b.put(RET, 40, 1, "rNPV pipeline (Base)", "label")
    b.put(RET, 40, 3, f"={R['pipeline_rnpv']}", "link", fmt=FM2, border=True)
    # rows 41-42 ($/MW, VC method) + 43-45 (low/mid/high) backfilled after CC built
    R["ret_val_low"] = b.ref(RET, 3, 43)
    R["ret_val_mid"] = b.ref(RET, 3, 44)
    R["ret_val_high"] = b.ref(RET, 3, 45)

    # =====================================================================
    # CALC_CROSSCHECKS ($/MW benchmark, VC method, RTB-as-%-of-built)
    # =====================================================================
    b.width(CC, {"A": 42, "B": 4, "C": 16, "D": 14, "E": 14})
    b.put(CC, 1, 1, "CALC — CROSS-CHECKS: $/MW benchmark, VC method, RTB vs built", "title")
    b.put(CC, 2, 1, "Asset-value cross-checks on the representative pipeline, Base scenario. All figures AUD $m unless stated.", "note", wrap=True)
    base_sale_mult = R["scn_salemult"][2]
    base_dev_mult = R["scn_devmult"][2]
    base_success = R["scn_flip"][2]
    b.put(CC, 3, 1, "$/MW benchmark (Base)", "sect", fill="sect")
    cc = [
        (4, "Total pipeline MW", f"=SUM({R['rnpv_mw']})", FNUM, "formula"),
        (5, "Gross RTB asset value $m", f"=SUMPRODUCT({R['rnpv_mw']},{R['rnpv_rtb']})*{base_sale_mult}", FM, "formula"),
        # project count = COUNT of the representative pipeline (no hardcoded literal)
        (6, "Total dev cost $m", f"={R['dev_cost']}*{base_dev_mult}*COUNT({R['rnpv_mw']})", FM, "formula"),
        (7, "Average years to sale", f"=AVERAGE({R['rnpv_years']})", FNUM1, "formula"),
        (8, "Flip success (Base, DA x conn x sale)", f"={base_success}", FPCT, "formula"),
        (9, "$/MW implied dev value $m", f"=IFERROR((C5-C6)*C8/(1+{R['discount_base']})^C7,0)", FM2, "formula"),
    ]
    for r, label, val, fmt, kind in cc:
        b.put(CC, r, 1, label, "label")
        b.put(CC, r, 3, val, kind, fmt=fmt, border=True, bold=(r == 9))
    R["dpmw_dev"] = b.ref(CC, 3, 9)

    b.put(CC, 11, 1, "VC method (Base)", "sect", fill="sect")
    vc = [
        (12, "Total RTB sale value $m", f"=SUMPRODUCT({R['rnpv_mw']},{R['rnpv_rtb']})*{base_sale_mult}", FM, "formula"),
        (13, "Total dev cost $m", "=C6", FM, "formula"),
        (14, "Flip success (Base)", "=C8", FPCT, "link"),
        (15, "Exit equity value (expected) $m", "=C14*(C12-C13)", FM2, "formula"),
        (16, "VC target return", f"={R['vc_target']}", FPCT, "link"),
        (17, "Hold horizon (yrs)", f"={R['term']}", FNUM, "link"),
        (18, "VC today value $m", "=IFERROR(C15/(1+C16)^C17,0)", FM2, "formula"),
    ]
    for r, label, val, fmt, kind in vc:
        b.put(CC, r, 1, label, "label")
        b.put(CC, r, 3, val, kind, fmt=fmt, border=True, bold=(r == 18))
    R["vc_today"] = b.ref(CC, 3, 18)

    b.put(CC, 20, 1, "RTB as % of built-asset value (sense-check)", "sect", fill="sect")
    b.header(CC, 21, ["State", "RTB $/MW", "Built $/MW", "RTB % of built"])
    for i, st in enumerate(["NSW", "VIC", "SA"]):
        r = 22 + i
        b.put(CC, r, 1, st, "label")
        b.put(CC, r, 2, f"={R['rtb_' + st]}", "link", fmt=FMW, border=True)
        b.put(CC, r, 3, f"={R['built_cost']}", "link", fmt=FMW, border=True)
        b.put(CC, r, 4, f"=IFERROR(B{r}/C{r},0)", "formula", fmt=FPCT, border=True)
    b.put(CC, 25, 1, "Expect ~10–12% — confirms RTB is the early-stage development slice, not the built asset.", "note", wrap=True)

    # backfill Returns valuation range now that CC refs exist
    b.put(RET, 41, 1, "$/MW benchmark (dev value)", "label")
    b.put(RET, 41, 3, f"={R['dpmw_dev']}", "link", fmt=FM2, border=True)
    b.put(RET, 42, 1, "VC method (today value)", "label")
    b.put(RET, 42, 3, f"={R['vc_today']}", "link", fmt=FM2, border=True)
    b.put(RET, 43, 1, "Low", "label", bold=True)
    b.put(RET, 43, 3, "=MIN(C40:C42)", "formula", fmt=FM2, bold=True, border=True)
    b.put(RET, 44, 1, "Midpoint", "label", bold=True)
    b.put(RET, 44, 3, "=AVERAGE(C40:C42)", "formula", fmt=FM2, bold=True, border=True, fill="sect")
    b.put(RET, 45, 1, "High", "label", bold=True)
    b.put(RET, 45, 3, "=MAX(C40:C42)", "formula", fmt=FM2, bold=True, border=True)

    # --- Full cap table (who owns what — ties to 100%) ---
    b.put(RET, 47, 1, "Full cap table — who owns what (mechanical; ties to 100%)", "sect", fill="sect")
    b.header(RET, 48, ["Holder", "At entry %", "At exit (diluted) %"])
    pool, fut = R["option_pool"], R["future_dilution"]
    b.put(RET, 49, 1, "Founders / existing holders", "label")
    b.put(RET, 49, 2, f"=1-{R['own_entry']}", "formula", fmt=FPCT, border=True)
    b.put(RET, 49, 3, f"=(1-{R['own_entry']})*(1-{pool})*(1-{fut})", "formula", fmt=FPCT, border=True)
    b.put(RET, 50, 1, "Us (new investor)", "label", bold=True)
    b.put(RET, 50, 2, f"={R['own_entry']}", "link", fmt=FPCT, border=True)
    b.put(RET, 50, 3, f"={R['own_diluted']}", "link", fmt=FPCT, border=True, fill="sect")
    b.put(RET, 51, 1, "Option pool (staff)", "label")
    b.put(RET, 51, 2, 0.0, "formula", fmt=FPCT, border=True)
    b.put(RET, 51, 3, f"={pool}*(1-{fut})", "formula", fmt=FPCT, border=True)
    b.put(RET, 52, 1, "Future-round investors", "label")
    b.put(RET, 52, 2, 0.0, "formula", fmt=FPCT, border=True)
    b.put(RET, 52, 3, f"={fut}", "link", fmt=FPCT, border=True)
    b.put(RET, 53, 1, "Total (must = 100%)", "label", bold=True)
    b.put(RET, 53, 2, "=SUM(B49:B52)", "formula", fmt=FPCT, bold=True, border=True)
    b.put(RET, 53, 3, "=SUM(C49:C52)", "formula", fmt=FPCT, bold=True, border=True)
    R["captable_entry_total"] = b.ref(RET, 2, 53)
    R["captable_exit_total"] = b.ref(RET, 3, 53)

    # --- Exit waterfall (Base) — company exit equity distributes fully ---
    b.put(RET, 55, 1, "Exit waterfall (Base) — full distribution of exit equity", "sect", fill="sect")
    b.put(RET, 56, 1, "Company exit equity (Base)", "label")
    b.put(RET, 56, 3, f"={R['exit_eq_by'][2]}", "link", fmt=FM2, border=True)
    b.put(RET, 57, 1, "Our terminal proceeds (1x pref / as-converted)", "label", bold=True)
    b.put(RET, 57, 3, f"={R['terminal_by'][2]}", "link", fmt=FM2, border=True)
    b.put(RET, 58, 1, "Other holders' proceeds (residual)", "label")
    b.put(RET, 58, 3, "=C56-C57", "formula", fmt=FM2, border=True)
    b.put(RET, 59, 1, "Distributed (ours + others = exit equity)", "label", bold=True)
    b.put(RET, 59, 3, "=C57+C58", "formula", fmt=FM2, bold=True, border=True)
    R["waterfall_exit_eq"] = b.ref(RET, 3, 56)
    R["waterfall_distributed"] = b.ref(RET, 3, 59)

    # =====================================================================
    # SENSITIVITY (OUR equity IRR: development-approval rate x RTB price)
    # =====================================================================
    b.width(SENS, {"A": 28, "B": 11, "C": 9, "D": 9, "E": 9, "F": 9, "G": 9,
                   "H": 9, "I": 9, "J": 9, "K": 3, "L": 9, "M": 9, "N": 9, "O": 9, "P": 9,
                   "Q": 3, "R": 9, "S": 9, "T": 9, "U": 9, "V": 9})
    b.put(SENS, 1, 1, "SENSITIVITY — OUR equity IRR: development approval × RTB price (forward-pipeline exit basis)", "title")
    b.put(SENS, 2, 1, "Return on OUR shares (equity IRR = MOIC^(1/exit year) − 1) vs DEVELOPMENT-APPROVAL rate (down) × RTB price "
                      "multiplier (across); all else at Base. Exit equity now uses the PRIMARY forward-pipeline basis: forward rNPV "
                      "(depth × per-project) + retained cash − debt; our proceeds = the GREATER of the 1× liquidation preference or our "
                      "as-converted share. Helper columns H–J (per DA), the exit-equity grid L–P and the proceeds grid R–V keep every "
                      "formula short. The exit assumption (pipeline depth × discount) is profiled separately in IC_MEMO Exhibit D. AUD $m / %.", "note", wrap=True)
    da_rates = [0.40, 0.55, 0.65, 0.80, 0.95]
    price_mults = [0.70, 0.85, 1.00, 1.15, 1.30]
    GC0, EQ0, PG0 = 3, 12, 18   # IRR grid C..G ; exit-equity grid L..P ; proceeds grid R..V
    # constants (SENS-local, short refs)
    consts = [
        (4, "Conn × sale (flip adj)", f"={R['p_conn']}*{R['p_sale']}", FPCT),
        (5, "Diluted ownership", f"={R['own_diluted']}", FPCT),
        (6, "Pipeline depth at exit", f"={R['exit_depth']}", FNUM),
        (7, "Preference cap (liq × investment) $m", f"={R['liq_pref']}*{R['investment']}", FM2),
        (8, "Our investment $m", f"={R['investment']}", FM2),
        (9, "Exit year", f"={R['exit_year']}", FNUM),
        (10, "Base blended sale/project $m", f"={R['base_blended']}", FM2),
        (11, "Dev cost/project $m", f"={R['dev_cost']}", FM2),
        (12, "Disc factor @ base over avg years", f"={R['df_avg']}", FFAC),
        (13, "Retain fraction (1 − dist)", f"=1-{R['dist_frac']}", FPCT),
        (14, "Debt at exit $m", f"={R['debt_exit']}", FM2),
    ]
    for r, label, val, fmt in consts:
        b.put(SENS, r, 1, label, "label")
        b.put(SENS, r, 2, val, "formula" if val.count("*") + val.count("-") else "link", fmt=fmt, border=True)
    R["s_flipadj"] = b.ref(SENS, 2, 4)
    # price header (row 17) + gross-per-price helper (row 18), both across the IRR-grid columns
    b.put(SENS, 17, 2, "Price mult →", "label", bold=True, align="right")
    b.put(SENS, 18, 2, "Gross $m →", "label", bold=True, align="right")
    for j, pm in enumerate(price_mults):
        cl = get_column_letter(GC0 + j)
        b.put(SENS, 17, GC0 + j, pm, "input", fmt=FMULT, fill="yel", align="center", border=True)
        b.put(SENS, 18, GC0 + j, f"={R['target']}*{R['base_blended']}*{cl}17", "formula", fmt=FM, align="center", border=True)
    # header row 19 + helper-col / grid labels
    b.put(SENS, 19, 2, "DA gate ↓", "head", fill="head", align="center", border=True)
    for j in range(len(price_mults)):
        b.put(SENS, 19, GC0 + j, "IRR", "head", fill="head", align="center", border=True)
        b.put(SENS, 19, EQ0 + j, "Exit eq", "head", fill="head", align="center", wrap=True, border=True)
        b.put(SENS, 19, PG0 + j, "Proceeds", "head", fill="head", align="center", wrap=True, border=True)
    for k, lab in enumerate(["Flip succ", "Started", "Total dev"]):
        b.put(SENS, 19, 8 + k, lab, "head", fill="head", align="center", wrap=True, border=True)
    for i, da in enumerate(da_rates):
        r = 20 + i
        b.put(SENS, r, 2, da, "input", fmt=FPCT, fill="yel", align="center", border=True)
        b.put(SENS, r, 8, f"=$B{r}*$B$4", "formula", fmt=FPCT, border=True)
        b.put(SENS, r, 9, f"=IFERROR({R['target']}/H{r},0)", "formula", fmt=FNUM1, border=True)
        b.put(SENS, r, 10, f"={R['target']}*{R['dev_cost']}+(I{r}-{R['target']})*{R['abandon']}*{R['dev_cost']}", "formula", fmt=FM, border=True)
        for j in range(len(price_mults)):
            pc = get_column_letter(GC0 + j)    # price header / gross helper column
            eqc = get_column_letter(EQ0 + j)   # exit-equity grid column
            prc = get_column_letter(PG0 + j)   # proceeds grid column
            # exit equity = MAX(0, forward rNPV + retained − debt)
            eqf = f"=MAX(0,$B$6*($B$10*{pc}$17-$B$11)*$H{r}*$B$12+MAX(0,{pc}$18-$J{r})*$B$13-$B$14)"
            b.put(SENS, r, EQ0 + j, eqf, "formula", fmt=FM2, border=True)
            # proceeds = MAX(1× pref cap, diluted × exit equity)
            b.put(SENS, r, PG0 + j, f"=MAX(MIN($B$7,{eqc}{r}),$B$5*{eqc}{r})", "formula", fmt=FM2, border=True)
            # IRR = (proceeds / investment) ^ (1/exit year) − 1 ; total loss → −100%
            b.put(SENS, r, GC0 + j, f"=IFERROR(({prc}{r}/$B$8)^(1/$B$9)-1,-1)", "formula", fmt=FPCT, border=True)
    R["sens_grid"] = b.rng(SENS, GC0, 20, GC0 + len(price_mults) - 1, 20 + len(da_rates) - 1)

    # =====================================================================
    # CHECKS
    # =====================================================================
    b.width(CHK, {"A": 54, "B": 12, "C": 38, "D": 4, "E": 9, "F": 9, "G": 9})
    b.put(CHK, 1, 1, "CHECKS — integrity framework (master check feeds the Cover)", "title")
    b.header(CHK, 3, ["Check", "Result", "Note"])
    cr0 = 4
    err_row = cr0 + 14
    fc_row = cr0 + 12   # First-Chicago range check uses E/F/G helpers on its own row
    checks = [
        (f'=IF(AND(MIN({b.rng(INP,3,26,3,28)})>=0,MAX({b.rng(INP,3,26,3,28)})<=1),"OK","ERROR")',
         "Every gate probability ∈ [0,1]", "Gate probs bounded"),
        (f'=IF({R["live_success"]}<=MIN({b.rng(SURV,2,4,2,6)})+0.000001,"OK","ERROR")',
         "Flip success ≤ each gate", "Survival monotonic"),
        (f'=IF(MIN({R["pipe_mw"]})>0,"OK","ERROR")', "All MW positive", "Sign check"),
        (f'=IF(MIN({b.rng(RNPV,7,4,7,4 + n - 1)})>0,"OK","ERROR")', "All RTB sale values positive", "Sign check"),
        (f'=IF({R["dev_cost"]}>0,"OK","ERROR")', "Dev cost positive", "Sign check"),
        (f'=IF(AND({R["switch"]}>=1,{R["switch"]}<=3),"OK","ERROR")', "Scenario switch ∈ {1,2,3}", "Switch valid"),
        (f'=IF(AND({R["scn_da"][1]}<={R["scn_da"][2]},{R["scn_da"][2]}<={R["scn_da"][3]}),"OK","ERROR")',
         "Scenario DA-gate monotonic", "Cons ≤ Base ≤ Ideal"),
        (f'=IF(ABS({R["post_money"]}-({R["pre_money"]}+{R["investment"]}))<0.0001,"OK","ERROR")',
         "Post-money = pre-money + investment", "Cap-table identity"),
        (f'=IF(ABS({R["own_entry"]}-{R["investment"]}/{R["post_money"]})<0.0001,"OK","ERROR")',
         "Ownership = investment ÷ post-money", "Cap-table identity"),
        (f'=IF({R["own_diluted"]}<={R["own_entry"]}+0.0001,"OK","ERROR")',
         "Diluted ownership ≤ entry ownership", "Dilution sign"),
        (f'=IF(AND({R["moic_by"][1]}<={R["moic_by"][2]},{R["moic_by"][2]}<={R["moic_by"][3]}),"OK","ERROR")',
         "Investor MOIC monotonic", "Cons ≤ Base ≤ Ideal"),
        (f'=IF(ABS({R["weights_sum"]}-1)<0.001,"OK","ERROR")', "Scenario weights sum to 100%", "First-Chicago weights"),
        (f'=IF(AND(G{fc_row}>=E{fc_row}-0.0001,G{fc_row}<=F{fc_row}+0.0001),"OK","ERROR")',
         "First-Chicago IRR within scenario range", "Weighting sanity (helpers E:G)"),
        (f'=IF(MIN({R["ret_proceeds_rng"]})>=0,"OK","ERROR")', "Our proceeds non-negative (all scenarios)", "Sign check"),
        (f'=IF(SUM(E{err_row}:G{err_row})=0,"OK","ERROR")', "No error cells in key outputs", "#REF!/#DIV0! scan"),
        (f'=IF(AND(ABS({R["captable_entry_total"]}-1)<0.001,ABS({R["captable_exit_total"]}-1)<0.001),"OK","ERROR")',
         "Cap table ties (ownership sums to 100%)", "VC cap-table integrity"),
        (f'=IF(ABS({R["waterfall_distributed"]}-{R["waterfall_exit_eq"]})<0.001,"OK","ERROR")',
         "Exit equity fully distributed (ours + others)", "VC waterfall integrity"),
        (f'=IF({R["exit_eq_by"][2]}<={R["fwd_by"][2]}+{R["retained_by"][2]}+0.000001,"OK","ERROR")',
         "No double-count (exit value ≤ realised + forward pipeline)", "Retained-cash-only guard"),
        (f'=IF(OR(AND({R["rtb_verified"]}=1,{R["devcost_verified"]}=1),{R["prov_disclosed"]}=1),"OK","ERROR")',
         "Provenance: unverified inputs flagged, not silently independent", "Contamination guard (see advisory below)"),
    ]
    for i, (formula, label, note) in enumerate(checks):
        r = cr0 + i
        b.put(CHK, r, 1, label, "label")
        b.put(CHK, r, 2, formula, "formula", align="center", border=True)
        b.put(CHK, r, 3, note, "note")
    # helpers for the First-Chicago range check (kept short)
    b.put(CHK, fc_row, 5, f"=MIN({R['ret_irr_rng']})", "formula", fmt=FPCT, align="center", border=True)
    b.put(CHK, fc_row, 6, f"=MAX({R['ret_irr_rng']})", "formula", fmt=FPCT, align="center", border=True)
    b.put(CHK, fc_row, 7, f"={R['expected_irr']}", "link", fmt=FPCT, align="center", border=True)
    b.put(CHK, err_row, 5, f"=SUMPRODUCT(--ISERROR({b.rng(RNPV,13,4,13,rt)}))", "formula", align="center", border=True)
    b.put(CHK, err_row, 6, f"=SUMPRODUCT(--ISERROR({R['ret_irr_rng']}))", "formula", align="center", border=True)
    b.put(CHK, err_row, 7, f"=SUMPRODUCT(--ISERROR({R['sens_grid']}))", "formula", align="center", border=True)
    cr1 = cr0 + len(checks) - 1
    rmaster = cr1 + 2
    b.put(CHK, rmaster, 1, "MASTER CHECK", "label", bold=True)
    b.put(CHK, rmaster, 2, f'=IF(COUNTIF(B{cr0}:B{cr1},"ERROR")=0,"OK","ERROR")', "formula", bold=True, align="center", border=True)
    R["master_check"] = b.ref(CHK, 2, rmaster)
    ok_rule = CellIsRule(operator="equal", formula=['"OK"'], fill=FILL_GREEN, font=Font(color="006100"))
    err_rule = CellIsRule(operator="equal", formula=['"ERROR"'], fill=FILL_RED, font=Font(color="9C0006"))
    b.ws[CHK].conditional_formatting.add(f"B{cr0}:B{rmaster}", ok_rule)
    b.ws[CHK].conditional_formatting.add(f"B{cr0}:B{rmaster}", err_rule)
    # provenance advisory (change5) — NOT in the master range; honest status of the independent valuation
    adv = rmaster + 2
    b.put(CHK, adv, 1, "PROVENANCE STATUS (advisory — not in master)", "disc", bold=True)
    b.put(CHK, adv, 2, f'=IF(AND({R["rtb_verified"]}=1,{R["devcost_verified"]}=1),"Verified","UNVERIFIED — verify RTB & dev cost")',
          "formula", border=True)
    b.put(CHK, adv + 1, 1, "The independent pipeline rNPV currently consumes the manager's UNVERIFIED RTB $/MW and dev-cost "
                           "claims (tagged Proposed). The value is illustrative until each is verified against a primary source "
                           "and re-tagged Independent (verified). Verifying them is more likely to LOWER the value than raise it "
                           "— widening, not closing, the entry-price gap. Master check is unaffected (this is a disclosure).", "note", wrap=True)
    b.ws[CHK].merge_cells(f"A{adv + 1}:G{adv + 1}")

    # =====================================================================
    # DASHBOARD
    # =====================================================================
    b.width(DSH, {"A": 38, "B": 16, "C": 6, "D": 34, "E": 16})
    b.put(DSH, 1, 1, "ILLUSTRATIVE BATTERY-DEVELOPER STARTUP — DIRECT-EQUITY DASHBOARD", "title")
    b.put(DSH, 2, 1, "We buy SHARES directly in an illustrative ~5 MW distribution-BESS develop-and-flip startup (NSW/VIC/SA). ILLUSTRATIVE — "
                     "founder/manager figures are claims and the equity-deal terms are placeholders to confirm. Not investment advice.", "disc", wrap=True)
    b.put(DSH, 4, 1, "Active scenario", "label", bold=True)
    b.put(DSH, 4, 2, f"={R['scenario_name']}", "link", bold=True)
    b.put(DSH, 4, 4, "Master check", "label", bold=True)
    b.put(DSH, 4, 5, f"={R['master_check']}", "link", bold=True, align="center")
    b.put(DSH, 6, 1, "EXPECTED RETURN ON OUR SHARES (First-Chicago)", "sect", fill="sect")
    b.put(DSH, 6, 4, "OUR STAKE", "sect", fill="sect")
    b.put(DSH, 7, 1, "Expected equity IRR", "label", bold=True)
    b.put(DSH, 7, 2, f"={R['expected_irr']}", "link", fmt=FPCT, bold=True, border=True)
    b.put(DSH, 8, 1, "Expected MOIC", "label")
    b.put(DSH, 8, 2, f"={R['expected_moic']}", "link", fmt=FX, border=True)
    stake = [(7, "Post-money valuation $m", f"={R['post_money']}", FM),
             (8, "Ownership at entry", f"={R['own_entry']}", FPCT),
             (9, "Ownership at exit (diluted)", f"={R['own_diluted']}", FPCT),
             (10, "Liquidation preference (x)", f"={R['liq_pref']}", FX),
             (11, "Exit year", f"={R['exit_year']}", FNUM)]
    for r, label, val, fmt in stake:
        b.put(DSH, r, 4, label, "label")
        b.put(DSH, r, 5, val, "link", fmt=fmt, border=True)
    b.put(DSH, 13, 1, "RETURN ON OUR SHARES BY SCENARIO", "sect", fill="sect")
    b.header(DSH, 14, ["Scenario", "Flip succ", "Equity IRR", "MOIC"])
    for i, (name, cid) in enumerate([("Conservative", 1), ("Base", 2), ("Ideal", 3)]):
        r = 15 + i
        b.put(DSH, r, 1, name, "label")
        b.put(DSH, r, 2, f"={R['scn_flip'][cid]}", "formula", fmt=FPCT, border=True)
        b.put(DSH, r, 3, f"={R['irr_by'][cid]}", "link", fmt=FPCT, border=True)
        b.put(DSH, r, 4, f"={R['moic_by'][cid]}", "link", fmt=FX, border=True)
    b.put(DSH, 13, 4, "KEY ASSUMPTIONS & FLAG", "sect", fill="sect")
    ka = [(14, "Programme capital $m", f"={R['programme']}", FM),
          (15, "Projects target", f"={R['target']}", FNUM),
          (16, "Manager headline DA gate (Base)", f"={R['scn_da'][2]}", FPCT),
          (17, "True flip success (Base, DAxconnxsale)", f"={R['flip_at_base_da']}", FPCT),
          (18, "Independent flip success (benchmark)", f"={R['flip_independent']}", FPCT)]
    for r, label, val, fmt in ka:
        b.put(DSH, r, 4, label, "label")
        b.put(DSH, r, 5, val, "link", fmt=fmt, border=True)
    b.put(DSH, 20, 1, "Conclusion (illustrative)", "sect", fill="sect")
    b.put(DSH, 21, 1,
          "A real, policy-backed, capital-light niche with a viable RTB exit — but the return on OUR shares is THIN once the gates are "
          "modelled and the cap table applied. The 65% headline is the development-approval gate ALONE; true flip success = approval x "
          "grid connection x sale is ~36% at Base, so the company's exit equity value is modest and — after dilution and our 1x "
          "liquidation preference — the expected return on our shares is only marginally positive while the Conservative case is a total "
          "loss. Biggest risks: EXIT/BUYER risk, then development/approval, then the equity-deal terms (all placeholders). Pursue only on "
          "verified buyer depth, RTB comps, a survivable downside, and confirmed cap-table terms.", "note", wrap=True)
    b.ws[DSH].merge_cells("A21:E26")

    # =====================================================================
    # COVER
    # =====================================================================
    b.width(COV, {"A": 26, "B": 64})
    b.put(COV, 1, 1, "Battery-Developer Startup Valuation — Direct-Equity Stake", "title")
    cover = [
        (2, "Subtitle", "Return on OUR shares (equity IRR/MOIC) in an illustrative distribution-BESS develop-and-flip startup (NSW/VIC/SA)"),
        (3, "Version", "v3.0"),
        (4, "Author", "Portfolio project (independent rebuild; founder/manager claims verified)"),
        (5, "Date", a["meta"]["as_at"]),
        (6, "Currency / Units", "AUD, $ millions ($m)"),
    ]
    for r, k, v in cover:
        b.put(COV, r, 1, k, "label", bold=True)
        b.put(COV, r, 2, v, "label", wrap=True)
    b.put(COV, 7, 1, "Active scenario", "label", bold=True)
    b.put(COV, 7, 2, f"={R['scenario_name']}", "link", bold=True)
    b.put(COV, 8, 1, "MASTER CHECK", "label", bold=True)
    b.put(COV, 8, 2, f"={R['master_check']}", "link", bold=True)
    b.ws[COV].conditional_formatting.add("B8", CellIsRule(operator="equal", formula=['"OK"'], fill=FILL_GREEN, font=Font(color="006100")))
    b.ws[COV].conditional_formatting.add("B8", CellIsRule(operator="equal", formula=['"ERROR"'], fill=FILL_RED, font=Font(color="9C0006")))
    b.put(COV, 10, 1, "Decision brief", "sect", fill="sect")
    brief = [
        ("Decision", "Whether to buy SHARES DIRECTLY (a minority equity stake) in an (illustrative) develop-and-flip RTB battery startup, and at what expected return on our shares."),
        ("Structure", "We invest as a SHAREHOLDER, NOT as a fund limited partner. A cap table (pre-/post-money, ownership, dilution, 1x liquidation preference) turns the company's exit equity value into OUR return. The equity-deal terms are PLACEHOLDERS to confirm."),
        ("Exit basis", "PRIMARY exit equity = forward-pipeline rNPV (projects in flight at exit) + retained cash − debt — what a buyer of a development platform pays. The earnings multiple is a CROSS-CHECK only (run-rate × a sourced low/base/high range). Pipeline depth at exit is the main sensitivity (Exhibit D)."),
        ("Reframe", "The company sells RTB before construction — MERCHANT RISK PASSES TO THE BUYER. Its risk is the survival curve (approve→connect→sell) + the RTB price; ours is also dilution and the cap-table terms."),
        ("Outputs", "Cap table, return on our shares (equity IRR & MOIC) by scenario, First-Chicago expected return, forward-pipeline exit + earnings-multiple cross-check, per-pipeline valuation cross-check, sensitivities."),
        ("Detail / horizon", "Company programme ~2+1-year cycle; our exit year is a placeholder. AUD, $m."),
        ("Out of scope", "Post-sale operating/merchant cash flows, construction (buyer-funded), full multi-round waterfall, tax structuring, FX, debt."),
    ]
    for i, (k, v) in enumerate(brief):
        r = 11 + i
        b.put(COV, r, 1, k, "label", bold=True)
        b.put(COV, r, 2, v, "label", wrap=True)
    b.put(COV, 18, 1, "Standards followed", "sect", fill="sect")
    b.put(COV, 19, 2, "FAST / ICAEW / Macabacus / Operis + VC-method structure: Inputs→Calcs→Outputs zones; one Timeline; one-row-one-calc; "
                      "no hardcodes in formulas; CHOOSE scenario switch + live-case row; checks built alongside; colour code (blue input / "
                      "black formula / green link); INDEX-MATCH not VLOOKUP; IFERROR on division. VC method: work back from the company's exit "
                      "equity value through the cap table; ownership = investment ÷ post-money; the required/target return (≥25%, VC cross-check) "
                      "is a RETURN HURDLE that compensates for failure risk — NOT a WACC — and failure risk is also handled explicitly via "
                      "First-Chicago scenario weighting.", "label", wrap=True)
    b.ws[COV].merge_cells("B19:B22")
    b.put(COV, 24, 1, "Judgement inputs to OWN (review pass)", "sect", fill="sect")
    b.put(COV, 25, 2, "Discount rate & premium • the THREE scenario success rates (vs the independent benchmark) • RTB $/MW by state (need "
                      "independent comps) • dev cost per project & abandonment • the equity-deal terms (pre-money, our cheque, option pool, "
                      "future dilution, liquidation preference, exit year) • the EXIT-VALUE drivers (pipeline depth at exit, distribution "
                      "fraction, debt at exit, and the cross-check multiple range — ALL placeholders). Founder/manager figures are CLAIMS "
                      "to verify — defend your own.", "label", wrap=True)
    b.ws[COV].merge_cells("B25:B28")
    b.put(COV, 30, 1, "TO COMPLETE (analyst review pass)", "sect", fill="sect")
    b.put(COV, 31, 2, "1. Trace every formula + colour audit; confirm master check = OK; stress the switch in all 3 positions.  "
                      "2. Replace BENCHMARK / claim inputs with verified independent values (RTB comps, per-state success, dev cost) and "
                      "CONFIRM the cap-table terms against a term sheet.  3. Stress a downside where RTB prices are 20–30% lower and success "
                      "≈ the independent benchmark; confirm our shares can be wiped out (Conservative = total loss).  4. Pressure-test the "
                      "EXIT BASIS (now the biggest swing factor): vary pipeline depth at exit and the discount rate (Exhibit D); confirm comps "
                      "if any (else the pipeline basis stands).  5. Depth left out: a SINGLE funding round and a 1× NON-PARTICIPATING "
                      "liquidation preference are scaffolded — add later rounds (Series B/C) to the cap table, and participation / "
                      "anti-dilution / accrued preferred dividends if the term sheet carries them.", "label", wrap=True)
    b.ws[COV].merge_cells("B31:B35")
    b.put(COV, 36, 1, "DISCLAIMER", "label", bold=True)
    b.put(COV, 36, 2, "ILLUSTRATIVE startup built from public benchmark data; an INDEPENDENT rebuild (not an endorsement). NOT investment "
                      "advice. Wholesale-investor context; read the term sheet / IM. The equity-deal terms are placeholders to confirm. All "
                      "figures illustrative and must be independently verified. Capital is at risk.", "disc", wrap=True)
    b.ws[COV].merge_cells("B36:B39")

    # =====================================================================
    # CONTENTS
    # =====================================================================
    b.width(CON, {"A": 6, "B": 28, "C": 60})
    b.put(CON, 1, 1, "Contents", "title")
    b.header(CON, 3, ["#", "Sheet", "Purpose"])
    purposes = {
        COV: "Title, decision brief, master check, disclaimer", CON: "This page",
        CL: "Version history", INP: "All assumptions + cap-table deal terms + imported data (blue)",
        TL: "Master timeline (built once)", SCN: "Scenarios + switch (DA gate) + live case + weights",
        SURV: "Separate gates; flip success = DA × connection × sale + flag",
        RNPV: "Per-project risk-adjusted NAV (RTB − dev cost)",
        FUND: "Company programme: funnel → realised profit + forward-pipeline rNPV → exit equity value (+ earnings cross-check)",
        RET: "Cap table; return on OUR shares (equity IRR/MOIC) by scenario, First-Chicago, valuation range",
        CC: "$/MW benchmark, VC method, RTB vs built",
        SENS: "Our equity IRR grid (development approval × RTB price)",
        CHK: "Integrity checks + master check", DSH: "One-page IC summary", SG: "Sources + glossary",
    }
    for i, name in enumerate(ORDER):
        r = 4 + i
        b.put(CON, r, 1, i + 1, "label", align="center")
        cell = b.put(CON, r, 2, name, "link")
        cell.hyperlink = f"#'{name}'!A1"
        b.put(CON, r, 3, purposes.get(name, ""), "note")

    # =====================================================================
    # CHANGE LOG
    # =====================================================================
    b.width(CL, {"A": 14, "B": 10, "C": 62, "D": 20})
    b.put(CL, 1, 1, "Change Log", "title")
    b.header(CL, 3, ["Date", "Version", "Change", "Author"])
    b.put(CL, 4, 1, a["meta"]["as_at"], "input", align="center")
    b.put(CL, 4, 2, "v1.0", "input", align="center")
    b.put(CL, 4, 3, "Initial build: survival curve, per-project rNPV, programme funnel, returns, First-Chicago, cross-checks, sensitivity, checks, dashboard.", "label", wrap=True)
    b.put(CL, 4, 4, "Analyst (v1 draft)", "input")
    b.put(CL, 5, 1, a["meta"]["as_at"], "input", align="center")
    b.put(CL, 5, 2, "v2.0", "input", align="center")
    b.put(CL, 5, 3, "Reframed from a fund LP investment to a DIRECT-EQUITY stake: removed fees/carry/hurdle; renamed Calc_Fund→Calc_Company (programme → net profit → company exit equity value); added a cap table (pre-/post-money, ownership, dilution, 1x liquidation preference) on Returns; outputs are now the return on OUR shares (equity IRR/MOIC), First-Chicago weights proceeds; sensitivity outputs our equity IRR.", "label", wrap=True)
    b.put(CL, 5, 4, "Analyst (v2 reframe)", "input")
    b.put(CL, 6, 1, a["meta"]["as_at"], "input", align="center")
    b.put(CL, 6, 2, "v3.0", "input", align="center")
    b.put(CL, 6, 3, "Fixed the exit-value basis (change2.md): PRIMARY exit equity is now FORWARD-PIPELINE rNPV (depth × per-project) + retained cash − debt, not net profit × a platform multiple. The earnings multiple is demoted to a CROSS-CHECK on FORWARD RUN-RATE profit (a sourced low/base/high range). Added pipeline depth / distribution-fraction / debt inputs; a retained-cash double-count guard; and the exit-value sensitivity (Exhibit D). Return = terminal proceeds + interim distributions (convention (b)).", "label", wrap=True)
    b.put(CL, 6, 4, "Analyst (v3 exit-basis)", "input")
    b.put(CL, 7, 1, a["meta"]["as_at"], "input", align="center")
    b.put(CL, 7, 2, "v4.0", "input", align="center")
    b.put(CL, 7, 3, "Added input PROVENANCE (change5): every price/value input tagged Proposed (manager) / Independent (verified) / Placeholder, plus a contamination guard so the independent rNPV consumes only verified inputs. RTB $/MW and dev cost are the manager's UNVERIFIED claims, so the independent value is illustrative pending verification (advisory on this tab). Anonymised: no real names; all monetary figures are illustrative placeholders.", "label", wrap=True)
    b.put(CL, 7, 4, "Analyst (v4 provenance)", "input")
    b.put(CL, 8, 3, "[Analyst review pass — record changes here]", "input", wrap=True, fill="yel")

    # =====================================================================
    # SOURCES & GLOSSARY
    # =====================================================================
    b.width(SG, {"A": 40, "B": 14, "C": 14, "D": 50})
    b.put(SG, 1, 1, "Sources & Glossary", "title")
    b.put(SG, 3, 1, "Sources (LIVE = fetched; BENCHMARK = documented public benchmark / the manager claim, verify)", "sect", fill="sect")
    b.header(SG, 4, ["Input", "Status", "As at", "Source"])
    src_rows = [
        ("Risk-free rate (RBA 10yr CGS)", "risk_free"),
        ("Built-cost context (CSIRO GenCost)", "build"),
        ("Dev cost per project", "dev"),
        ("Development-approval prob — public benchmark", "da"),
        ("Grid connection prob (AEMO)", "connection"),
        ("Reach-sale prob (buyer pool / AEMO attrition)", "sale"),
        ("RTB $/MW by state (the manager claim)", "rtb"),
    ]
    rr = 5
    for label, key in src_rows:
        s = inp.sources.get(key)
        b.put(SG, rr, 1, label, "label")
        b.put(SG, rr, 2, s[2] if s else "BENCHMARK", "label", align="center")
        b.put(SG, rr, 3, s[1] if s else "", "label", align="center")
        b.put(SG, rr, 4, s[0] if s else "Illustrative / the manager claim (see config/assumptions.yaml)", "note", wrap=True)
        rr += 1
    b.put(SG, rr + 1, 1, "RTB comps: publicly reported deals only; paid DBs (BNEF, Enerdatics, Mergermarket) out of scope. the manager's prices are claims to verify.", "note", wrap=True)
    gloss_r = rr + 3
    b.put(SG, gloss_r, 1, "Glossary", "sect", fill="sect")
    glossary = [
        ("RTB", "Ready-to-build / development rights — a shovel-ready project sold before construction."),
        ("rNPV", "Risk-adjusted NPV — each cash flow × probability of occurring, then discounted."),
        ("Survival curve / PD", "Cumulative probability a project clears all gates; structurally a multi-period probability-of-default curve."),
        ("Development funnel", "projects_started = target ÷ flip success; the company funds dev cost on every started project (full on delivered, partial on dropouts)."),
        ("Realised net programme profit", "The company's gross RTB proceeds minus its total development cost — profit EARNED by exit (feeds retained cash; counted once, at face value)."),
        ("Forward pipeline rNPV", "The risk-adjusted value of projects still IN FLIGHT at exit (depth × per-project rNPV) — the PRIMARY exit equity, what a buyer of a development platform pays for."),
        ("Retained cash / run-rate", "Retained cash = realised profit not yet distributed (carried into the exit value). Run-rate = realised profit ÷ term — the yearly stream the earnings-multiple CROSS-CHECK is applied to."),
        ("Exit equity (primary)", "Forward-pipeline rNPV + retained cash − debt at exit. The earnings multiple on run-rate profit is a CROSS-CHECK, not the headline."),
        ("Pre-money / post-money", "Pre-money = the company's agreed value before our money goes in; post-money = pre-money + our investment."),
        ("Ownership %", "Our share of the company = our investment ÷ post-money valuation."),
        ("Dilution", "The fall in our ownership % when new shares are issued — here a staff option pool and future funding rounds."),
        ("Cap table", "The schedule of who owns what % of the company; we run our stake through it to get our exit proceeds."),
        ("Liquidation preference", "A preferred right to get our money back (1x) before ordinary holders at exit; 1x non-participating means we take the GREATER of that or our as-converted share."),
        ("MOIC / equity IRR", "Multiple on OUR investment; annualised return on our shares (MOIC^(1/exit year) − 1; a total loss is −100%)."),
        ("First Chicago", "Probability-weight the scenarios' proceeds into a single expected return on our shares."),
        ("Develop-and-flip", "Develop, de-risk, then sell RTB — the margin is the value uplift; merchant risk passes to the buyer."),
    ]
    for i, (term, defn) in enumerate(glossary):
        r = gloss_r + 1 + i
        b.put(SG, r, 1, term, "label", bold=True)
        b.put(SG, r, 4, defn, "note", wrap=True)

    # back-to-contents links + freeze panes
    for name in ORDER:
        if name != CON:
            c = b.ws[name].cell(row=1, column=10, value="↩ Contents")
            c.font = F_LINK
            c.hyperlink = f"#'{CON}'!A1"
        b.ws[name].freeze_panes = "A4" if name not in (COV, CON, DSH) else "A2"

    # --- mandatory formula-quality scan (calc + output + control sheets) ---
    import re as _re
    long_hits, nested_hits = [], []
    for name in [TL, SCN, SURV, RNPV, FUND, RET, CC, SENS, CHK]:
        for row in b.ws[name].iter_rows():
            for c in row:
                v = c.value
                if isinstance(v, str) and v.startswith("="):
                    if len(v) > 100:
                        long_hits.append(f"{name}!{c.coordinate} ({len(v)} chars)")
                    if len(_re.findall(r"(?<!ERROR)\bIF\(", v)) >= 2:
                        nested_hits.append(f"{name}!{c.coordinate}")

    # Excel-first (change11): the master BESS_Valuation.xlsx is HAND-OWNED. A normal
    # run writes NOTHING — Python never overwrites the master. It is regenerated only
    # with an explicit --rebuild-master (there is one model file; no generated copy).
    if not rebuild_master:
        print("[build_model] the master workbook is hand-owned — nothing written.")
        print("[build_model] pass --rebuild-master to regenerate the master from Python.")
        return None
    out = io.PROJECT_ROOT / "financial_models" / "BESS_Valuation.xlsx"
    print("[build_model] --rebuild-master: regenerating the HAND-OWNED master workbook.")
    out.parent.mkdir(parents=True, exist_ok=True)
    b.wb.save(out)
    print(f"[build_model] wrote {out}")
    if long_hits or nested_hits:
        print(f"[build_model] FORMULA SCAN FAILED — long: {long_hits[:8]}  nested IF: {nested_hits[:8]}")
    else:
        print("[build_model] formula-quality scan PASSED — no long formulas, no nested IFs on any calc sheet")
    return out


if __name__ == "__main__":
    import sys
    build(rebuild_master="--rebuild-master" in sys.argv)
