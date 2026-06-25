"""Extractor: CSIRO GenCost battery capital cost -> data/processed/costs.csv

GenCost ships a multi-sheet .xlsx whose layout changes each release, so we:
  1. discover the current workbook link on the GenCost landing page,
  2. download it (cached, date-stamped),
  3. best-effort scan for a battery $/kWh figure,
  4. else fall back to the documented benchmark (status BENCHMARK).

Output costs.csv columns: item, value, unit, source, as_at, status
"""
from __future__ import annotations

import csv

from src.utils import io
from src.utils import sources_log


def _try_parse_battery_cost(xlsx_path) -> float | None:
    """Scan the workbook for a plausible battery storage $/kWh (returns $m/MWh)."""
    try:
        from openpyxl import load_workbook

        wb = load_workbook(xlsx_path, read_only=True, data_only=True)
        for ws in wb.worksheets:
            for row in ws.iter_rows(values_only=True):
                cells = [str(c).lower() if c is not None else "" for c in row]
                line = " ".join(cells)
                if "battery" in line and ("kwh" in line or "storage" in line):
                    for c in row:
                        if isinstance(c, (int, float)) and 150 <= float(c) <= 1200:
                            # $/kWh -> $m per MWh: ($/kWh * 1000 kWh/MWh) / 1e6
                            return round(float(c) * 1000 / 1e6, 4)
        return None
    except Exception as exc:  # noqa: BLE001
        print(f"  [parse-fail] GenCost workbook: {exc}")
        return None


def run() -> None:
    print("[extract] CSIRO GenCost battery cost (build-cost CONTEXT only — the buyer funds construction)")
    src = io.load_sources()["sources"]["csiro_gencost"]
    assum = io.load_assumptions()
    dev_cfg = assum["dev_cost"]
    built_cfg = assum["built_cost_context"]

    build_per_mwh = None
    status = "BENCHMARK"

    link = io.discover_file_link(src["landing_page"], src["discover_link_pattern"])
    if link:
        dest = io.DATA_RAW / f"csiro_gencost_{io.today_str()}.xlsx"
        path = io.cached_download(link, dest, binary=True)
        if path:
            build_per_mwh = _try_parse_battery_cost(path)
            if build_per_mwh is not None:
                status = "LIVE"

    # Built-cost CONTEXT per MW (sense-check for 'RTB as % of built value'). If
    # GenCost gave a live $/MWh, convert at a 2-hour reference duration; else use
    # the documented benchmark. This is CONTEXT only — it is NOT in the fund's
    # margin (the fund sells RTB pre-construction; the buyer funds the build).
    if build_per_mwh is not None:
        built_per_mw = round(build_per_mwh * 2.0, 4)   # 2h reference duration
    else:
        io.manual_download_msg(src["name"], src["landing_page"],
                               io.DATA_RAW / "csiro_gencost_<date>.xlsx")
        built_per_mw = float(built_cfg["per_mw_m"]["value"])

    dev_per_project = float(dev_cfg["per_project_m"]["value"])
    out = io.DATA_PROCESSED / "costs.csv"
    with open(out, "w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow(["item", "value", "unit", "source", "as_at", "status"])
        w.writerow(["dev_cost_per_project_m", round(dev_per_project, 4), "$m per project",
                    dev_cfg["per_project_m"]["source"],
                    dev_cfg["per_project_m"]["as_at"], "BENCHMARK"])
        w.writerow(["built_cost_context_per_mw_m", built_per_mw, "$m per MW (context)",
                    built_cfg["per_mw_m"]["source"],
                    built_cfg["per_mw_m"]["as_at"], status])
    print(f"  -> wrote {out} (dev={dev_per_project} $m/project [BENCHMARK], "
          f"built-context={built_per_mw} $m/MW [{status}])")

    sources_log.record(
        "csiro_gencost",
        name=src["name"],
        url=src["landing_page"],
        output="data/processed/costs.csv",
        status=status,
        description=src["description"],
    )


if __name__ == "__main__":
    run()
