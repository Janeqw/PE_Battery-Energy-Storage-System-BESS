"""Extractor: AEMO NEM Generation Information.

Two jobs:
  1. Ground the illustrative pipeline — list NSW/VIC batteries (size/location)
     written to data/raw for clean_pipeline.py to reference.
  2. Measure proposed -> committed attrition as a proxy for the 'reach sale'
     gate probability, staged for gate_statistics.py.

Verify-then-fallback: discover & download the current workbook; if unreachable,
fall back to the documented benchmark for the sale gate.
"""
from __future__ import annotations

import csv

from src.utils import io
from src.utils import sources_log

_BATTERY_HINTS = ("battery", "bess", "storage")
_STATUS_COMMITTED = ("committed", "commissioning", "existing", "operational", "operating")
_STATUS_PROPOSED = ("proposed", "publicly announced", "advanced", "anticipated")


def _parse_geninfo(xlsx_path):
    """Return (nsw_vic_batteries:list[dict], n_proposed:int, n_committed:int)."""
    try:
        from openpyxl import load_workbook

        wb = load_workbook(xlsx_path, read_only=True, data_only=True)
        batteries, n_prop, n_comm = [], 0, 0
        for ws in wb.worksheets:
            rows = ws.iter_rows(values_only=True)
            header = None
            for row in rows:
                cells = [str(c).strip() if c is not None else "" for c in row]
                low = [c.lower() for c in cells]
                if header is None:
                    if any("region" in c for c in low) and any("fuel" in c or "technology" in c for c in low):
                        header = low
                    continue
                line = " ".join(low)
                if not any(h in line for h in _BATTERY_HINTS):
                    continue
                is_nsw_vic = (" nsw" in f" {line}") or ("vic" in line)
                if any(s in line for s in _STATUS_COMMITTED):
                    n_comm += 1
                elif any(s in line for s in _STATUS_PROPOSED):
                    n_prop += 1
                if is_nsw_vic:
                    batteries.append({"raw": " | ".join(cells)})
        return batteries, n_prop, n_comm
    except Exception as exc:  # noqa: BLE001
        print(f"  [parse-fail] AEMO GenInfo: {exc}")
        return [], 0, 0


def run() -> None:
    print("[extract] AEMO Generation Information")
    src = io.load_sources()["sources"]["aemo_generation_info"]
    sale_cfg = io.load_assumptions()["gates"]["reach_sale"]

    sale_prob = None
    status = "BENCHMARK"
    n_prop = n_comm = 0

    link = io.discover_file_link(src["landing_page"], src["discover_link_pattern"])
    if link:
        dest = io.DATA_RAW / f"aemo_generation_information_{io.today_str()}.xlsx"
        path = io.cached_download(link, dest, binary=True)
        if path:
            batteries, n_prop, n_comm = _parse_geninfo(path)
            if (n_prop + n_comm) >= 10:  # enough signal to compute attrition
                sale_prob = round(n_comm / (n_prop + n_comm), 3)
                status = "LIVE"
                # write the grounding list of NSW/VIC batteries
                with open(io.DATA_RAW / "aemo_nswvic_batteries.csv", "w",
                          newline="", encoding="utf-8") as fh:
                    w = csv.writer(fh)
                    w.writerow(["raw_row"])
                    for b in batteries:
                        w.writerow([b["raw"]])

    if sale_prob is None:
        io.manual_download_msg(src["name"], src["landing_page"],
                               io.DATA_RAW / "aemo_generation_information_<date>.xlsx")
        sale_prob = float(sale_cfg["probability"])

    io.write_stage("sale", {
        "probability": sale_prob,
        "duration_years": float(sale_cfg["duration_years"]),
        "status": status,
        "n_proposed": n_prop,
        "n_committed": n_comm,
        "source": sale_cfg["source"],
        "as_at": sale_cfg["as_at"],
    })
    print(f"  -> sale gate p={sale_prob} [{status}] (proposed={n_prop}, committed={n_comm})")

    sources_log.record(
        "aemo_generation_info",
        name=src["name"],
        url=src["landing_page"],
        output="data/processed/pipeline.csv, gate_stats.csv (sale gate)",
        status=status,
        description=src["description"],
    )


if __name__ == "__main__":
    run()
