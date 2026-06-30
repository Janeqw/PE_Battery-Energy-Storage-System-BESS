"""Refresh the Excel model's BLUE INPUT cells from the processed CSVs / engine.

This is the data-layer fallback to Power Query: it writes input VALUES into the
workbook's input cells only — it never touches a formula cell (any cell whose
value starts with '=' is skipped).

Excel-first: the master BESS_Valuation.xlsx is HAND-OWNED. A normal run writes NOTHING.
To push refreshed inputs into the master's input cells (formulas never touched), pass
--write-master explicitly.

    python -m src.refresh_model_inputs                 # no-op (master is hand-owned)
    python -m src.refresh_model_inputs --write-master  # updates the master's input cells

If the master does not exist yet, it is built first (with --write-master).
"""
from __future__ import annotations

from openpyxl import load_workbook

from src.utils import io
from src.valuation_engine import load_inputs
from src import build_model

MASTER = io.PROJECT_ROOT / "financial_models" / "BESS_Valuation.xlsx"

# label substring (on Inputs col A) -> attribute on the Inputs dataclass (col C)
SCALAR_MAP = {
    "Risk-free rate": "risk_free",
    "Development risk premium": "risk_premium",
    "Programme capital": "programme_capital",
    "Projects target": "projects_target",
    "Term": "term_years",
    "Development cost per project": "dev_cost_per_project",
    "Abandonment fraction": "abandonment_fraction",
    "Built-asset cost": "built_cost_per_mw",
    "Development approval": "da_independent",
    "Grid connection": "p_connection",
    "Reach sale": "p_sale",
    "Pre-money valuation": "pre_money",
    "Our investment": "investment_amount",
    "Option pool": "option_pool_pct",
    "Future-round dilution": "future_round_dilution_pct",
    "Liquidation preference": "liquidation_pref_x",
    "Exit year": "exit_year",
    "Pipeline depth at exit": "pipeline_depth_at_exit",
    "Interim distribution fraction": "interim_distribution_fraction",
    "Debt at exit": "debt_at_exit",
    "Cross-check multiple — low": "xmult_low",
    "Cross-check multiple — base": "xmult_base",
    "Cross-check multiple — high": "xmult_high",
    "VC target return": "vc_target_return",
}
DURATION_MAP = {"Development approval": "dur_da", "Grid connection": "dur_connection", "Reach sale": "dur_sale"}
STATES = ("NSW", "VIC", "SA")  # exact match on col A -> rtb_comps[state]


def _set_input(ws, row, col, value):
    """Write to an input cell only if it isn't a formula."""
    cell = ws.cell(row=row, column=col)
    if isinstance(cell.value, str) and cell.value.startswith("="):
        return False
    cell.value = value
    return True


def run(write_master: bool = False) -> None:
    if not write_master:
        print("[refresh] the master workbook is hand-owned — nothing written.")
        print("[refresh] pass --write-master to push refreshed inputs into the master.")
        return
    target = MASTER
    print("[refresh] --write-master: updating input cells in the HAND-OWNED master "
          "(formulas untouched).")
    if not target.exists():
        print("[refresh] master not found — building it first")
        build_model.build(rebuild_master=True)

    inp = load_inputs()
    wb = load_workbook(target)
    ws = wb["Inputs"]
    n_updated = 0

    for row in ws.iter_rows(min_col=1, max_col=1):
        cell = row[0]
        label = str(cell.value or "").strip()
        r = cell.row
        for key, attr in SCALAR_MAP.items():
            if label.startswith(key):
                if _set_input(ws, r, 3, getattr(inp, attr)):
                    n_updated += 1
                if key in DURATION_MAP:
                    _set_input(ws, r, 8, getattr(inp, DURATION_MAP[key]))
        if label in STATES:  # RTB $/MW by state
            if _set_input(ws, r, 3, float(inp.rtb_comps[label])):
                n_updated += 1
        if label.startswith("Capital call profile"):
            for j, v in enumerate(inp.call_profile[:4]):
                _set_input(ws, r, 3 + j, float(v))
            n_updated += 1
        if label.startswith("Distribution profile"):
            for j, v in enumerate(inp.dist_profile[:4]):
                _set_input(ws, r, 3 + j, float(v))
            n_updated += 1

    # Pipeline table: find header row "Project", write project rows beneath it.
    header_row = None
    for row in ws.iter_rows(min_col=1, max_col=1):
        if str(row[0].value or "").strip() == "Project":
            header_row = row[0].row
            break
    if header_row:
        for i, p in enumerate(inp.projects):
            r = header_row + 1 + i
            _set_input(ws, r, 1, p["name"])
            _set_input(ws, r, 2, p["state"])
            _set_input(ws, r, 3, p["location"])
            _set_input(ws, r, 4, p["mw"])
            _set_input(ws, r, 5, p["duration_h"])
            _set_input(ws, r, 7, p["years_to_sale"])
            n_updated += 1

    wb.calculation.fullCalcOnLoad = True
    wb.save(target)
    print(f"[refresh] updated {n_updated} input cells in {target.name} (formulas untouched)")
    print("          Open in Excel to recalculate (fullCalcOnLoad is set).")


if __name__ == "__main__":
    import sys
    run(write_master="--write-master" in sys.argv)
