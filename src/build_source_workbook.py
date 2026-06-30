"""Generate financial_models/INPUTS_AND_ASSUMPTIONS.xlsx — the single SOURCE workbook
the model (BESS_Valuation.xlsx) links to via Power Query (change15).

Two data sheets (+ a README sheet):
  * Inputs       — SOURCED data (from data/raw/), each figure showing where it comes
                   from (source name, source_key, raw file, page/table or URL, date,
                   provenance). Built from the verification trail + config/sources.yaml.
  * Assumptions  — JUDGEMENT assumptions, rendered from config/assumptions.yaml
                   (source_key: none), each showing how it's derived + any [[TO CONFIRM]].

This file is GENERATED — do not hand-edit. Inputs come from data/raw/, assumptions
from config/assumptions.yaml. Rebuild:  python -m src.build_source_workbook

One unit per column; one number per cell (committee-report rules). No fabricated
pages — figures without a pinned page carry [[TO CONFIRM: PDF + page]].
"""
from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from src.utils import io

# ---------------------------------------------------------------------------
# Styles (financial-modelling skill: blue = input/data, headers shaded)
# ---------------------------------------------------------------------------
BLUE, BLACK, GREY = "0000FF", "000000", "595959"
F_TITLE = Font(name="Arial", size=14, bold=True, color="1F3864")
F_HEAD = Font(name="Arial", size=10, bold=True, color="FFFFFF")
F_SECT = Font(name="Arial", size=10, bold=True, color="1F3864")
F_DATA = Font(name="Arial", size=10, color=BLUE)       # sourced/judgement values = inputs
F_LABEL = Font(name="Arial", size=10, color=BLACK)
F_NOTE = Font(name="Arial", size=9, italic=True, color=GREY)
FILL_HEAD = PatternFill("solid", fgColor="1F3864")
FILL_SECT = PatternFill("solid", fgColor="D9E1F2")
FILL_GEN = PatternFill("solid", fgColor="FFF2CC")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")


# ---------------------------------------------------------------------------
# Inputs sheet — SOURCED data (the verification trail, structured).
# No fabricated pages: where a page isn't pinned, the locator says so.
# columns: Item, Value, Unit, Source name, source_key, Raw file, Page/Table or URL, Date, Provenance
# ---------------------------------------------------------------------------
AEMO_ISP_URL = "https://www.aemo.com.au/energy-systems/major-publications/integrated-system-plan-isp/2026-integrated-system-plan-isp"
SOURCED_INPUTS = [
    ("RBA — Capital Market Yields (F2)", [
        ("10-year CGS yield", 4.776, "%", "rba", "data/raw/rba-cgs-10yr-2026-06.csv",
         "col 'Australian Government 10 year bond', row 17-Jun-2026", "2026-06-17", "Independent (verified)"),
    ]),
    ("CSIRO — GenCost 2025-26", [
        ("2-hour battery build cost (cell)", 525, "$/kWh", "csiro_gencost", "data/raw/csiro-gencost-2025-26.pdf",
         "current capital-cost tables, 2-hr battery [[TO CONFIRM: page]]", "2025-12", "Independent (verified)"),
    ]),
    ("AEMO — 2026 Integrated System Plan", [
        ("Small-scale storage power (2030)", 12, "GW", "aemo_isp", "data/raw/aemo-2026-isp.pdf",
         "Storage capacity, Step Change [[TO CONFIRM: page]]", "2026", "Independent (verified)"),
        ("Small-scale storage energy (2030)", 33, "GWh", "aemo_isp", "data/raw/aemo-2026-isp.pdf",
         "Storage capacity, Step Change [[TO CONFIRM: page]]", "2026", "Independent (verified)"),
        ("Grid-scale storage need (2050)", 40, "GW", "aemo_isp", "data/raw/aemo-2026-isp.pdf",
         "Storage capacity, Step Change (35 + 5) [[TO CONFIRM: page]]", "2026", "Independent (verified)"),
        ("Grid-scale storage in connections", 45, "GW", "aemo_isp", "data/raw/aemo-2026-isp.pdf",
         "Connections / transition outlook (~45 of ~67 GW) [[TO CONFIRM: page]]", "2026", "Independent (verified)"),
    ]),
    ("AEMO — gate benchmarks (Connections / Generation Information)", [
        ("Grid-connection rate (RTB rights: agreement + GPS)", 70, "%", "aemo_connections",
         "data/raw/ (Scorecard not pinned)", "registration / executed connection agreement stage; NOT energisation",
         "2026-06", "[[TO CONFIRM]] benchmark"),
        ("Reach-sale rate (proposed->committed attrition)", 80, "%", "aemo_generation_info",
         "data/raw/aemo_generation_information_<date>.xlsx", "proposed -> committed status",
         "2026-06", "[[TO CONFIRM]] benchmark"),
    ]),
    ("NSW Planning Portal — approval benchmark", [
        ("Development-approval rate (independent benchmark)", 80, "%", "nsw_planning",
         "data/raw/nsw_planning_<date>.json", "approved vs refused/withdrawn BESS major projects",
         "2026-06", "[[TO CONFIRM]] benchmark"),
    ]),
    ("Trade press — RTB $/MW comps (manager claim to verify)", [
        ("RTB sale price — NSW", 0.20, "$m/MW", "trade_press_comps", "data/raw/trade_press_candidates_20260624.csv",
         "trade-press deal comps (independent comps needed)", "2026-06", "Proposed (manager)"),
        ("RTB sale price — VIC", 0.18, "$m/MW", "trade_press_comps", "data/raw/trade_press_candidates_20260624.csv",
         "trade-press deal comps (independent comps needed)", "2026-06", "Proposed (manager)"),
        ("RTB sale price — SA", 0.12, "$m/MW", "trade_press_comps", "data/raw/trade_press_candidates_20260624.csv",
         "trade-press deal comps (independent comps needed)", "2026-06", "Proposed (manager)"),
        ("All-in built-asset value (context; above GenCost cell cost)", 1.80, "$m/MW", "csiro_gencost",
         "data/raw/csiro-gencost-2025-26.pdf", "all-in value ABOVE the ~$1.1m/MW cell cost — no primary",
         "2026-06", "Placeholder [[TO CONFIRM]]"),
    ]),
]
INPUT_COLS = ["Item", "Value", "Unit", "Source name", "source_key",
              "Raw file (data/raw/…)", "Page/Table or URL", "Date", "Provenance"]

# Theme grouping for the Assumptions sheet (top-level YAML key -> theme label).
ASSUMPTION_THEMES = [
    ("Discount / required return", ["discount_rate", "returns"]),
    ("Company programme", ["company"]),
    ("Equity deal terms", ["equity_deal"]),
    ("Exit drivers", ["exit_value", "exit_value_crosscheck"]),
    ("Development cost & attrition", ["dev_cost"]),
    ("Gates / scenarios (manager DA claim)", ["scenarios"]),
    ("Pipeline", ["pipeline"]),
    ("Value ladder", ["value_ladder"]),
]
ASSUMPTION_COLS = ["Item", "Value", "Unit", "How it's derived / rationale",
                   "source_key", "[[TO CONFIRM]]", "as_at"]


def _judgement_rows(section: dict, prefix: str = "") -> list:
    """Walk a YAML section; yield (item, value, rationale, as_at, to_confirm) for every
    node tagged source_key: none (a judgement assumption)."""
    rows = []
    if not isinstance(section, dict):
        return rows
    if section.get("source_key") == "none":
        val = section.get("value")
        if val is None:                         # block-level judgement (pipeline / scenarios)
            val = ""
        if isinstance(val, list):
            val = ", ".join(str(x) for x in val)
        src = str(section.get("source", ""))
        rows.append((prefix.rstrip("."), val, src, str(section.get("as_at", "")),
                     "Y" if "[[TO CONFIRM" in src else ""))
    for k, v in section.items():
        if isinstance(v, dict):
            rows.extend(_judgement_rows(v, f"{prefix}{k}."))
    return rows


def _style_header(ws, row, cols):
    for j, c in enumerate(cols, start=1):
        cell = ws.cell(row, j, c)
        cell.font = F_HEAD
        cell.fill = FILL_HEAD
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.border = BORDER


def _build_readme(wb, build_cmd):
    ws = wb.active
    ws.title = "README"
    ws.column_dimensions["A"].width = 110
    ws["A1"] = "INPUTS_AND_ASSUMPTIONS.xlsx — the model's SOURCE workbook"
    ws["A1"].font = F_TITLE
    lines = [
        "",
        "GENERATED — DO NOT HAND-EDIT.",
        "",
        "• Inputs sheet  = SOURCED data, built from data/raw/ (each figure shows source, raw file, page/URL, date, provenance).",
        "• Assumptions sheet = JUDGEMENT assumptions, rendered from config/assumptions.yaml (source_key: none).",
        "",
        "The model BESS_Valuation.xlsx links to these two sheets via Power Query (relative path). It types no source numbers.",
        "",
        "Change a SOURCED input  -> update the source data / re-run extractors, then rebuild this workbook.",
        "Change a JUDGEMENT assumption -> edit config/assumptions.yaml, then rebuild this workbook.",
        "",
        f"Rebuild:  {build_cmd}",
        "Config (human-readable, GitHub-inspectable): config/assumptions.yaml, config/sources.yaml, config/SOURCES_LOG.md",
    ]
    for i, t in enumerate(lines, start=2):
        cell = ws.cell(i, 1, t)
        cell.font = F_SECT if t.endswith("DO NOT HAND-EDIT.") else F_LABEL
        if t.startswith("Rebuild") or t.startswith("Config"):
            cell.font = F_NOTE
    ws["A3"].fill = FILL_GEN


def _build_inputs(wb):
    ws = wb.create_sheet("Inputs")
    widths = [44, 9, 8, 34, 20, 38, 46, 10, 22]
    for j, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + j) if j <= 26 else "A"].width = w
    ws["A1"] = "Inputs — SOURCED data (from data/raw/). GENERATED — do not hand-edit."
    ws["A1"].font = F_TITLE
    _style_header(ws, 3, INPUT_COLS)
    r = 4
    for group, rows in SOURCED_INPUTS:
        gc = ws.cell(r, 1, group); gc.font = F_SECT; gc.fill = FILL_SECT
        for j in range(2, len(INPUT_COLS) + 1):
            ws.cell(r, j).fill = FILL_SECT
        r += 1
        for item, val, unit, skey, raw, loc, date, prov in rows:
            ws.cell(r, 1, item).font = F_LABEL
            vcell = ws.cell(r, 2, val); vcell.font = F_DATA
            ws.cell(r, 3, unit).font = F_LABEL
            ws.cell(r, 4, _source_name(skey)).font = F_LABEL
            ws.cell(r, 5, skey).font = F_LABEL
            ws.cell(r, 6, raw).font = F_NOTE
            lc = ws.cell(r, 7, loc); lc.font = F_NOTE; lc.alignment = WRAP
            ws.cell(r, 8, date).font = F_LABEL
            ws.cell(r, 9, prov).font = F_LABEL
            for j in range(1, len(INPUT_COLS) + 1):
                ws.cell(r, j).border = BORDER
            r += 1
    ws.freeze_panes = "A4"


def _build_assumptions(wb, assum):
    ws = wb.create_sheet("Assumptions")
    widths = [40, 16, 12, 70, 16, 14, 12]
    for j, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + j)].width = w
    ws["A1"] = "Assumptions — JUDGEMENT (from config/assumptions.yaml). GENERATED — do not hand-edit."
    ws["A1"].font = F_TITLE
    _style_header(ws, 3, ASSUMPTION_COLS)
    r = 4
    for theme, keys in ASSUMPTION_THEMES:
        block = []
        for key in keys:
            block.extend(_judgement_rows(assum.get(key, {}), f"{key}."))
        if not block:
            continue
        gc = ws.cell(r, 1, theme); gc.font = F_SECT; gc.fill = FILL_SECT
        for j in range(2, len(ASSUMPTION_COLS) + 1):
            ws.cell(r, j).fill = FILL_SECT
        r += 1
        for item, val, rationale, as_at, tc in block:
            ws.cell(r, 1, item).font = F_LABEL
            vcell = ws.cell(r, 2, val); vcell.font = F_DATA
            ws.cell(r, 3, "").font = F_LABEL                      # unit: most are decimals/counts; left blank
            rc = ws.cell(r, 4, rationale); rc.font = F_NOTE; rc.alignment = WRAP
            ws.cell(r, 5, "none / judgement").font = F_LABEL
            ws.cell(r, 6, tc).font = F_LABEL
            ws.cell(r, 7, as_at).font = F_LABEL
            for j in range(1, len(ASSUMPTION_COLS) + 1):
                ws.cell(r, j).border = BORDER
            r += 1
    ws.freeze_panes = "A4"


_SOURCE_NAMES = {}


def _source_name(skey: str) -> str:
    return _SOURCE_NAMES.get(skey, skey)


def build() -> "Path":
    global _SOURCE_NAMES
    srcs = io.load_sources()["sources"]
    _SOURCE_NAMES = {k: v.get("name", k) for k, v in srcs.items() if k != "context_only"}
    assum = io.load_assumptions()

    wb = Workbook()
    build_cmd = "python -m src.build_source_workbook"
    _build_readme(wb, build_cmd)
    _build_inputs(wb)
    _build_assumptions(wb, assum)

    out = io.PROJECT_ROOT / "financial_models" / "INPUTS_AND_ASSUMPTIONS.xlsx"
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    print(f"[build_source_workbook] wrote {out} (sheets: {wb.sheetnames})")
    return out


if __name__ == "__main__":
    build()
